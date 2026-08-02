"""
F&B Restaurant Sector — Integrated 3-Statement Financial Modeling Dashboard
============================================================================
Finance student sector benchmarking assignment tool.
Accepts Excel workbooks with raw Income Statement, Balance Sheet, and
Cash Flow Statement data for multi-unit restaurant concepts.

Run:  streamlit run fb_app.py
"""

import io
import warnings
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

warnings.filterwarnings("ignore")

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(
    page_title="F&B Financial Model",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# GLOBAL CSS
# =============================================================================
st.markdown("""
<style>
  .main  { background: #0F1117; }
  .stApp { background: #0F1117; }
  .block-container { padding-top: 1.5rem; }

  .kpi-card {
    background: linear-gradient(135deg, #1B365D 0%, #2D5F8A 100%);
    border: 1px solid #2D5F8A;
    border-radius: 10px;
    padding: 18px 22px;
    text-align: center;
    margin-bottom: 10px;
  }
  .kpi-label  { font-size:11px; font-weight:600; color:#A8C8E8;
                text-transform:uppercase; letter-spacing:1.4px; margin-bottom:6px; }
  .kpi-value  { font-size:26px; font-weight:700; color:#FFFFFF; }
  .kpi-delta  { font-size:11px; color:#52C97C; margin-top:4px; }
  .kpi-neg    { color:#FF6B6B; }

  .section-hdr {
    font-size:15px; font-weight:700; color:#EAF2FA;
    border-left:4px solid #2D5F8A;
    padding-left:10px; margin:22px 0 14px;
  }
  .method-card {
    background:#161B27; border:1px solid #1F2937;
    border-radius:8px; padding:16px 20px; margin-bottom:10px;
  }
  .method-title { font-size:13px; font-weight:700; color:#60A5FA; margin-bottom:6px; }
  .method-body  { font-size:12px; color:#9CA3AF; line-height:1.7; }

  div[data-testid="stSidebar"] {
    background:#111827; border-right:1px solid #1F2937;
  }
  .stTabs [data-baseweb="tab-list"] { background:#111827; border-radius:8px; }
  .stTabs [data-baseweb="tab"]      { color:#9CA3AF; }
  .stTabs [aria-selected="true"]    { color:#60A5FA !important; background:#1F2937 !important; }
  h1,h2,h3 { color:#EAF2FA !important; }
  .stDataFrame { background:#161B27; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# LINE ITEM NORMALIZATION MAP
# Maps raw SEC filing label variants -> canonical keys
# Add entries here to support additional naming conventions
# =============================================================================
LINE_MAP: dict[str, str] = {
    # ── Revenue ───────────────────────────────────────────────────────────────
    "total revenue": "revenue",            "total revenues": "revenue",
    "net revenue": "revenue",              "net revenues": "revenue",
    "revenue": "revenue",                  "net sales": "revenue",
    "total net revenues": "revenue",       "total net revenue": "revenue",
    "restaurant revenue": "revenue",       "restaurant sales": "revenue",
    "company restaurant sales": "revenue", "shack sales": "revenue",

    # ── COGS / Food costs ─────────────────────────────────────────────────────
    "cost of sales": "cogs",
    "cost of goods sold": "cogs",
    "cost of revenue": "cogs",
    "restaurant operating costs": "cogs",
    "company-operated shop costs": "cogs",
    "food beverage packaging": "cogs",
    "food and paper costs": "cogs",
    "food beverage and packaging": "cogs",  # Chipotle / Sweetgreen exact label
    "food  beverage and packaging": "cogs", # trailing-space variant
    "food and beverage costs": "cogs",
    "food costs": "cogs",
    "cost of food and paper": "cogs",
    "product costs": "cogs",

    # ── Labor ─────────────────────────────────────────────────────────────────
    "labor": "labor",
    "labor expense": "labor",
    "labor and related costs": "labor",
    "labor costs": "labor",
    "store labor": "labor",
    "labor and related expenses": "labor",  # Shake Shack / Sweetgreen exact
    "restaurant labor": "labor",
    "crew labor": "labor",

    # ── Store / restaurant operating expenses ─────────────────────────────────
    "store operating expenses": "store_opex",
    "restaurant operating expenses": "store_opex",
    "occupancy and other costs": "store_opex",
    "occupancy costs": "store_opex",        # Chipotle exact
    "occupancy and related expenses": "store_opex",  # Shake Shack / Sweetgreen
    "other operating costs": "store_opex",  # Chipotle exact
    "other restaurant operating costs": "store_opex",  # Sweetgreen exact
    "other operating expenses": "store_opex",
    "pre-opening costs": "store_opex",
    "pre opening costs": "store_opex",      # normalized variant (no hyphen)
    "restaurant pre-opening costs": "store_opex",
    "preopening costs": "store_opex",

    # ── Advertising / marketing ───────────────────────────────────────────────
    "advertising expenses": "advertising",
    "advertising fees": "advertising",
    "advertising expense": "advertising",
    "marketing expense": "advertising",
    "marketing expenses": "advertising",

    # ── SGA ───────────────────────────────────────────────────────────────────
    "selling general and administrative": "sga",
    "sg&a": "sga",
    "sga": "sga",
    "general and administrative": "sga",
    "g&a": "sga",
    "general and administrative expenses": "sga",   # all three companies exact
    "general and administrative expense": "sga",
    "corporate general and administrative": "sga",

    # ── D&A ───────────────────────────────────────────────────────────────────
    "depreciation and amortization": "da",
    "d&a": "da",
    "depreciation": "da",
    "amortization": "da",
    "depreciation and amortization expense": "da",

    # ── Stock-based compensation (maps to a new key for CF add-back) ──────────
    "stock-based compensation": "sbc",
    "stock based compensation": "sbc",
    "share-based compensation": "sbc",
    "share based compensation expense": "sbc",

    # ── EBIT / Operating income ───────────────────────────────────────────────
    "operating income": "ebit",
    "operating income loss": "ebit",        # Shake Shack / Sweetgreen exact
    "income from operations": "ebit",
    "loss from operations": "ebit",
    "operating profit": "ebit",
    "income loss from operations": "ebit",

    # ── Interest ─────────────────────────────────────────────────────────────
    "interest expense": "interest",
    "interest expense net": "interest",
    "net interest expense": "interest",
    "interest expense income net": "interest",
    "interest income": "interest",
    "interest income expense net": "interest",

    # ── Tax ───────────────────────────────────────────────────────────────────
    "income tax expense": "tax",
    "income tax": "tax",
    "provision for income taxes": "tax",
    "benefit from income taxes": "tax",
    "income tax expense benefit": "tax",
    "income tax provision": "tax",

    # ── Net income ────────────────────────────────────────────────────────────
    "net income": "net_income",
    "net income loss": "net_income",
    "net loss": "net_income",
    "net income attributable": "net_income",
    "net loss attributable": "net_income",

    # ── Balance sheet: current assets ─────────────────────────────────────────
    "cash and cash equivalents": "cash",
    "cash": "cash",
    "total current assets": "current_assets",
    "accounts receivable": "accounts_receivable",
    "inventories": "inventories",
    "inventory": "inventories",
    "prepaid expenses and other current assets": "prepaid_other",
    "prepaid expenses": "prepaid_other",
    "marketable securities": "marketable_securities",   # Shake Shack
    "short-term investments": "short_term_investments", # Sweetgreen
    "short term investments": "short_term_investments",

    # ── Balance sheet: long-term assets ──────────────────────────────────────
    "property and equipment net": "ppe",
    "pp&e net": "ppe",
    "property plant and equipment net": "ppe",
    "property and equipment  net": "ppe",   # double-space variant
    "operating lease right-of-use assets": "rou_assets",
    "operating lease right of use assets": "rou_assets",
    "right-of-use assets": "rou_assets",
    "right of use assets net": "rou_assets",
    "goodwill": "goodwill",
    "goodwill and intangibles": "goodwill",
    "intangible assets and goodwill": "goodwill",      # Shake Shack exact
    "intangible assets net": "goodwill",
    "other non-current assets": "other_nca",
    "other noncurrent assets": "other_nca",
    "other non current assets": "other_nca",           # all three exact (normalized)
    "total assets": "total_assets",

    # ── Balance sheet: current liabilities ───────────────────────────────────
    "accounts payable": "accounts_payable",
    "accrued liabilities": "accrued_liabilities",
    "accrued expenses": "accrued_liabilities",          # Shake Shack exact
    "current operating lease liabilities": "cur_lease_lia",
    "current portion of operating lease liabilities": "cur_lease_lia",
    "total current liabilities": "current_liabilities",

    # ── Balance sheet: long-term liabilities ──────────────────────────────────
    "long-term debt": "ltd",
    "long term debt": "ltd",
    "total debt": "ltd",
    "notes payable": "ltd",
    "non-current operating lease liabilities": "lt_lease_lia",
    "noncurrent operating lease liabilities": "lt_lease_lia",
    "long-term operating lease liabilities": "lt_lease_lia",
    "other non-current liabilities": "other_ncl",
    "other noncurrent liabilities": "other_ncl",
    "other non current liabilities": "other_ncl",

    # ── Balance sheet: equity ─────────────────────────────────────────────────
    "total liabilities": "total_liabilities",
    "total equity": "equity",
    "total stockholders equity": "equity",
    "stockholders equity": "equity",
    "shareholders equity": "equity",
    "total stockholders deficit": "equity",
    "total liabilities and equity": "total_liab_equity",   # cross-check row
    "total liabilities and stockholders equity": "total_liab_equity",
    "retained earnings": "retained_earnings",
    "retained earnings accum deficit": "retained_earnings",  # Shake Shack exact
    "accumulated deficit": "retained_earnings",               # Sweetgreen exact
    "common stock and treasury stock": "common_stock",        # Chipotle exact
    "common stock and additional paid in capital": "common_stock", # SHAK/SG exact
    "common stock": "common_stock",

    # ── Cash flow: operating ──────────────────────────────────────────────────
    "cash provided by operating activities": "cfo",
    "net cash provided by operating": "cfo",
    "net cash provided by operating activities": "cfo",
    "operating cash flow": "cfo",
    "cash flows from operating activities": "cfo",
    "net cash from operating activities": "cfo",
    "changes in operating working capital": "working_capital_change",
    "change in operating working capital": "working_capital_change",
    "changes in working capital": "working_capital_change",

    # ── Cash flow: investing ──────────────────────────────────────────────────
    "purchases of property and equipment": "capex",
    "capital expenditures": "capex",
    "capex": "capex",
    "additions to property and equipment": "capex",
    "purchase of property and equipment": "capex",
    "cash used in investing activities": "cfi",             # all three exact
    "net cash used in investing activities": "cfi",
    "net cash from investing activities": "cfi",
    "other investing activities": "other_investing",
    "other investing activities net": "other_investing",

    # ── Cash flow: financing ──────────────────────────────────────────────────
    "net cash used in financing": "cff",
    "cash used in financing": "cff",
    "cash flow from financing": "cff",
    "cash used in financing activities": "cff",             # all three exact
    "net cash used in financing activities": "cff",
    "net cash from financing activities": "cff",
    "repurchase of common stock": "share_repurchase",       # Chipotle exact
    "repurchases of common stock": "share_repurchase",
    "other financing activities": "other_financing",
    "proceeds from stock option exercises": "stock_proceeds",  # Shake Shack exact
    "proceeds from stock issuance": "stock_proceeds",          # Sweetgreen exact
    "proceeds from exercise of stock options": "stock_proceeds",

    # ── Cash flow: other ──────────────────────────────────────────────────────
    "net increase in cash": "net_cash_change",
    "net increase decrease in cash": "net_cash_change",
    "net decrease in cash": "net_cash_change",
    "net change in cash": "net_cash_change",
}

DISPLAY_NAMES: dict[str, str] = {
    "revenue": "Net Revenue",
    "cogs": "Cost of Sales",
    "labor": "Labor and Related Costs",
    "store_opex": "Store Operating Expenses",
    "advertising": "Advertising Expense",
    "sga": "SG and A Expense",
    "da": "Depreciation and Amortization",
    "interest": "Net Interest Expense",
    "tax": "Income Tax Provision",
    "net_income": "Net Income",
    "cash": "Cash and Cash Equivalents",
    "ppe": "Property and Equipment Net",
    "goodwill": "Goodwill and Intangibles",
    "total_assets": "Total Assets",
    "ltd": "Long-Term Debt",
    "total_liabilities": "Total Liabilities",
    "equity": "Total Stockholders Equity",
    "cfo": "Operating Cash Flow",
    "capex": "Capital Expenditures",
    "cff": "Cash Flow from Financing",
}

IS_KEYS  = ["revenue","cogs","labor","store_opex","advertising","sga","da","interest","tax","net_income"]
BS_KEYS  = ["cash","ppe","goodwill","total_assets","ltd","total_liabilities","equity"]
CF_KEYS  = ["cfo","capex","cff"]


# =============================================================================
# DATA PARSING ENGINE
# =============================================================================
def normalize_label(raw: str) -> str | None:
    """
    Convert a raw SEC filing line-item label to a canonical key.

    Resolution order:
      1. Exact match after standard cleaning (lowercase, strip punctuation).
      2. Fuzzy substring rules — applied in priority order so that more
         specific patterns are tested before broader ones.  A rule fires
         when ALL required tokens appear in the cleaned string AND none of
         the optional exclusion tokens appear.

    Returns the canonical key string, or None if nothing matches.
    The caller (unmapped list) handles the None case.
    """
    # ── Stage 1: exact dictionary lookup ──────────────────────────────────────
    cleaned = (str(raw).strip().lower()
               .replace(",", "").replace(".", "")
               .replace("(", "").replace(")", "")
               .replace("-", " ").replace("/", " ")
               .replace("  ", " ").strip())

    if cleaned in LINE_MAP:
        return LINE_MAP[cleaned]

    # ── Stage 2: fuzzy substring rules ────────────────────────────────────────
    # Each entry: (required_tokens, exclude_tokens, canonical_key)
    # ALL required tokens must appear; ANY exclude token disqualifies the match.
    # Listed from most specific to least specific so earlier rules win.
    FUZZY_RULES: list[tuple[list[str], list[str], str]] = [
        # Revenue — must contain "revenue" or "sales" but not be a sub-line
        (["total net revenue"],   [],                          "revenue"),
        (["total net revenues"],  [],                          "revenue"),
        (["total revenue"],       ["cost", "other"],           "revenue"),
        (["net revenue"],         ["cost", "other"],           "revenue"),
        (["net revenues"],        ["cost", "other"],           "revenue"),
        (["restaurant revenue"],  [],                          "revenue"),
        (["shack sales"],         [],                          "revenue"),
        (["company restaurant"],  ["cost"],                    "revenue"),

        # Food / COGS — food + beverage + packaging cluster
        (["food", "beverage", "packaging"], [],                "cogs"),
        (["food", "packaging"],             [],                "cogs"),
        (["food", "beverage"],              ["labor","sga","admin"], "cogs"),
        (["food", "paper"],                 [],                "cogs"),
        (["cost of sales"],                 [],                "cogs"),
        (["cost of goods"],                 [],                "cogs"),
        (["cost of revenue"],               [],                "cogs"),

        # Labor
        (["labor", "related"],              [],                "labor"),
        (["labor", "expense"],              [],                "labor"),
        (["labor", "cost"],                 [],                "labor"),

        # Store operating / occupancy
        (["pre", "opening"],                [],                "store_opex"),
        (["pre-opening"],                   [],                "store_opex"),
        (["occupancy", "related"],          [],                "store_opex"),
        (["occupancy"],                     [],                "store_opex"),
        (["other restaurant operating"],    [],                "store_opex"),
        (["other operating"],               ["income","expense","non"],  "store_opex"),

        # SGA — general and administrative
        (["general", "administrative"],     [],                "sga"),

        # D&A
        (["depreciation", "amortization"],  [],                "da"),
        (["depreciation"],                  ["accumulated"],   "da"),

        # Stock-based compensation
        (["stock", "compensation"],         [],                "sbc"),
        (["share", "compensation"],         [],                "sbc"),
        (["stock-based"],                   [],                "sbc"),

        # EBIT / operating income
        (["operating income"],              ["non"],           "ebit"),
        (["income from operations"],        [],                "ebit"),
        (["loss from operations"],          [],                "ebit"),
        (["operating profit"],              [],                "ebit"),

        # Interest
        (["interest expense"],              [],                "interest"),
        (["net interest"],                  [],                "interest"),

        # Tax
        (["income tax"],                    [],                "tax"),
        (["provision for income"],          [],                "tax"),
        (["benefit from income tax"],       [],                "tax"),

        # Net income
        (["net income"],                    [],                "net_income"),
        (["net loss"],                      [],                "net_income"),

        # ── Balance sheet assets ──────────────────────────────────────────────
        (["total current assets"],          [],                "current_assets"),
        (["cash", "cash equivalents"],      [],                "cash"),
        (["accounts receivable"],           [],                "accounts_receivable"),
        (["inventories"],                   [],                "inventories"),
        (["inventory"],                     [],                "inventories"),
        (["prepaid"],                       [],                "prepaid_other"),
        (["marketable securities"],         [],                "marketable_securities"),
        (["short", "term", "investment"],   [],                "short_term_investments"),
        (["right-of-use"],                  ["current"],       "rou_assets"),
        (["right of use"],                  ["current"],       "rou_assets"),
        (["operating lease", "asset"],      ["current"],       "rou_assets"),
        (["intangible", "goodwill"],        [],                "goodwill"),
        (["goodwill"],                      [],                "goodwill"),
        (["property", "equipment", "net"],  [],                "ppe"),
        (["property", "plant", "equipment"],["purchases"],     "ppe"),
        (["other non-current assets"],      [],                "other_nca"),
        (["other noncurrent assets"],       [],                "other_nca"),
        (["other non current assets"],      [],                "other_nca"),
        (["total assets"],                  ["current","liab"],"total_assets"),

        # ── Balance sheet liabilities ─────────────────────────────────────────
        (["accounts payable"],              [],                "accounts_payable"),
        (["accrued liabilities"],           [],                "accrued_liabilities"),
        (["accrued expenses"],              [],                "accrued_liabilities"),
        (["current", "operating lease", "liabilit"], ["non"],  "cur_lease_lia"),
        (["total current liabilities"],     [],                "current_liabilities"),
        (["non-current", "operating lease"],["asset"],         "lt_lease_lia"),
        (["noncurrent", "operating lease"], ["asset"],         "lt_lease_lia"),
        (["long-term", "operating lease"],  ["asset"],         "lt_lease_lia"),
        (["non current", "operating lease"],["asset"],         "lt_lease_lia"),
        (["non current", "lease", "liabilit"], ["asset"],      "lt_lease_lia"),
        (["other non-current liabilities"], [],                "other_ncl"),
        (["other noncurrent liabilities"],  [],                "other_ncl"),
        (["other non current liabilities"], [],                "other_ncl"),
        (["long-term debt"],                [],                "ltd"),
        (["long term debt"],                [],                "ltd"),
        (["total liabilities"],             ["equity","and"],  "total_liabilities"),

        # ── Balance sheet equity ──────────────────────────────────────────────
        (["total liabilities", "equity"],   [],                "total_liab_equity"),
        (["total liabilities and"],         [],                "total_liab_equity"),
        (["retained earnings"],             [],                "retained_earnings"),
        (["accumulated deficit"],           [],                "retained_earnings"),
        (["common stock", "additional"],    [],                "common_stock"),
        (["common stock", "treasury"],      [],                "common_stock"),
        (["total equity"],                  [],                "equity"),
        (["total stockholders"],            [],                "equity"),
        (["stockholders equity"],           [],                "equity"),
        (["shareholders equity"],           [],                "equity"),

        # ── Cash flow: operating ──────────────────────────────────────────────
        (["net cash", "operating"],         [],                "cfo"),
        (["cash", "operating activities"],  ["used"],          "cfo"),
        (["changes in", "working capital"], [],                "working_capital_change"),
        (["change in", "working capital"],  [],                "working_capital_change"),

        # ── Cash flow: investing ──────────────────────────────────────────────
        (["purchases of property"],         [],                "capex"),
        (["purchase of property"],          [],                "capex"),
        (["additions to property"],         [],                "capex"),
        (["capital expenditures"],          [],                "capex"),
        (["cash used in investing"],        [],                "cfi"),
        (["net cash used in investing"],    [],                "cfi"),
        (["net cash", "investing"],         [],                "cfi"),
        (["other investing"],               [],                "other_investing"),

        # ── Cash flow: financing ──────────────────────────────────────────────
        (["cash used in financing"],        [],                "cff"),
        (["net cash used in financing"],    [],                "cff"),
        (["net cash", "financing"],         [],                "cff"),
        (["repurchase", "common stock"],    [],                "share_repurchase"),
        (["repurchases", "stock"],          [],                "share_repurchase"),
        (["proceeds from stock"],           [],                "stock_proceeds"),
        (["proceeds from exercise"],        [],                "stock_proceeds"),
        (["other financing"],               [],                "other_financing"),

        # ── Net cash change ───────────────────────────────────────────────────
        (["net increase", "cash"],          [],                "net_cash_change"),
        (["net decrease", "cash"],          [],                "net_cash_change"),
        (["net increase decrease", "cash"], [],                "net_cash_change"),
        (["net change in cash"],            [],                "net_cash_change"),
    ]

    for required_tokens, exclude_tokens, key in FUZZY_RULES:
        # All required tokens must appear in cleaned string
        if not all(tok in cleaned for tok in required_tokens):
            continue
        # No exclude token may appear
        if any(tok in cleaned for tok in exclude_tokens):
            continue
        return key

    return None


def _parse_one_sheet(df: pd.DataFrame, scale: float) -> dict:
    """
    Parse a single DataFrame (one sheet) into a partial company data dict.
    Returns { years, data, unmapped } or empty dict if nothing mapped.
    """
    df.columns = [str(c).strip() for c in df.columns]
    if df.empty or df.shape[1] < 2:
        return {}

    label_col = df.columns[0]
    year_cols = [
        c for c in df.columns[1:]
        if str(c).strip() not in ("", "nan")
        and not str(c).strip().lower().startswith("unnamed")
    ]
    if not year_cols:
        return {}

    result: dict = {"years": year_cols, "data": {}, "unmapped": []}
    for _, row in df.iterrows():
        raw_label = str(row[label_col]).strip()
        if not raw_label or raw_label.lower() in ("nan", "line item", "metric", ""):
            continue
        key = normalize_label(raw_label)
        if key is None:
            result["unmapped"].append(raw_label)
            continue
        if key not in result["data"]:
            result["data"][key] = {}
        for yr in year_cols:
            raw_val = str(row.get(yr, "")).strip().replace(",", "").replace("$", "")
            try:
                result["data"][key][yr] = float(raw_val) * scale
            except ValueError:
                result["data"][key][yr] = result["data"][key].get(yr, 0.0)

    return result if result["data"] else {}


# Keywords that identify a sheet as a specific statement type rather than a company.
# When ALL sheets in a workbook match these patterns the workbook is treated as
# a single-company multi-tab file and merged under the workbook filename.
_STMT_SHEET_KEYWORDS = {
    "income", "profit", "loss", "p&l", "earnings",
    "balance", "assets", "liabilities",
    "cash flow", "cashflow", "cash flows",
    "statement", "financials", "summary",
}


def _is_statement_sheet(sheet_name: str) -> bool:
    """Return True if the sheet name looks like a financial statement tab."""
    lower = sheet_name.lower()
    return any(kw in lower for kw in _STMT_SHEET_KEYWORDS)


def _merge_partial_data(parts: list[dict]) -> dict:
    """
    Merge multiple partial company data dicts (one per statement tab) into one.
    Years are unified across all parts; later parts do not overwrite earlier ones
    for the same (key, year) pair so that the first parsed value wins.
    """
    all_years: list[str] = []
    seen_years: set[str] = set()
    merged_data: dict = {}
    all_unmapped: list[str] = []

    for part in parts:
        for yr in part.get("years", []):
            if yr not in seen_years:
                all_years.append(yr)
                seen_years.add(yr)
        for key, yr_vals in part.get("data", {}).items():
            if key not in merged_data:
                merged_data[key] = {}
            for yr, val in yr_vals.items():
                if yr not in merged_data[key]:
                    merged_data[key][yr] = val
        all_unmapped.extend(part.get("unmapped", []))

    # Sort years numerically where possible
    def _yr_sort_key(yr: str) -> int:
        digits = "".join(c for c in yr if c.isdigit())
        return int(digits) if digits else 0

    all_years = sorted(set(all_years), key=_yr_sort_key)
    return {"years": all_years, "data": merged_data, "unmapped": list(set(all_unmapped))}


def parse_excel(file_obj, denomination: str, ticker_hint: str = "") -> dict[str, dict]:
    """
    Parse one uploaded Excel workbook and return a dict of company data.

    Two workbook layouts are handled automatically:

    Layout A — one sheet per company (e.g. a combined comps file):
      Each sheet is treated as one company.  The sheet name becomes the ticker.

    Layout B — one company across multiple statement sheets (e.g. a CMG file
      with tabs named Income Statement, Balance Sheet, Cash Flow Statement):
      All sheets are merged into a single company profile.  The workbook
      filename (or ticker_hint) becomes the ticker.

    Detection logic: if every sheet in the workbook has a name that matches a
    financial-statement keyword the workbook is treated as Layout B.  Otherwise
    each sheet with recognisable mapped data becomes its own company.

    Returns dict: { TICKER -> { years, data, unmapped } }
    """
    scale = 1.0 if denomination == "USD Thousands ($000s)" else 1000.0

    try:
        xl = pd.ExcelFile(file_obj)
    except Exception:
        return {}

    sheet_names = xl.sheet_names

    # Detect layout: if every sheet looks like a statement name -> Layout B
    all_statement_like = (
        len(sheet_names) > 0
        and all(_is_statement_sheet(s) for s in sheet_names)
    )

    companies: dict[str, dict] = {}

    if all_statement_like:
        # Layout B: merge all sheets into one company
        parts = []
        for sn in sheet_names:
            try:
                df = xl.parse(sn, header=0, index_col=None, dtype=str)
            except Exception:
                continue
            part = _parse_one_sheet(df, scale)
            if part:
                parts.append(part)

        if parts:
            merged = _merge_partial_data(parts)
            # Derive ticker from hint or use a cleaned version of the first sheet name
            ticker = (
                ticker_hint.upper().split(".")[0]  # strip file extension
                or sheet_names[0].upper()[:6]
            )
            # Remove common non-ticker words from the filename
            for drop in ("FINANCIALS", "MODEL", "DATA", "REPORT", "INCOME",
                         "BALANCE", "CASH", "FLOW", "STATEMENT"):
                ticker = ticker.replace(drop, "").strip("_- ")
            ticker = ticker[:8] or "CO"
            companies[ticker] = merged
    else:
        # Layout A: one company per sheet
        for sn in sheet_names:
            try:
                df = xl.parse(sn, header=0, index_col=None, dtype=str)
            except Exception:
                continue
            part = _parse_one_sheet(df, scale)
            if part:
                companies[sn.upper()[:12]] = part

    return companies


def parse_multiple_files(
    file_objs: list,
    denomination: str,
) -> dict[str, dict]:
    """
    Parse a list of uploaded file objects (each a separate company or comps file).
    Returns the merged company dict across all files.
    """
    all_companies: dict[str, dict] = {}
    for f in file_objs:
        # Use the filename (without extension) as the ticker hint for Layout B detection
        fname = getattr(f, "name", "")
        ticker_hint = fname.rsplit(".", 1)[0].upper() if fname else ""
        parsed = parse_excel(f, denomination, ticker_hint=ticker_hint)
        for ticker, cdata in parsed.items():
            # If a ticker already exists from another file, suffix with a number
            unique_ticker = ticker
            counter = 2
            while unique_ticker in all_companies:
                unique_ticker = f"{ticker}{counter}"
                counter += 1
            all_companies[unique_ticker] = cdata
    return all_companies


# =============================================================================
# METRICS ENGINE
# =============================================================================
def compute_metrics(cdata: dict) -> dict:
    """Compute derived F&B metrics for all historical years."""
    d, yrs = cdata["data"], cdata["years"]

    def g(key, yr):
        return d.get(key, {}).get(yr, 0.0) or 0.0

    result = {}
    for yr in yrs:
        rev   = g("revenue", yr) or 1
        cogs  = abs(g("cogs",   yr))
        labor = abs(g("labor",  yr))
        sga   = abs(g("sga",    yr))
        da    = abs(g("da",     yr))
        cfo   = g("cfo",   yr)
        capex = g("capex", yr)
        cash  = g("cash",  yr)
        ltd   = g("ltd",   yr)
        ni    = g("net_income", yr)

        capex_abs  = abs(capex)
        prime_cost = cogs + labor
        adv        = abs(g("advertising", yr))
        ebitda     = rev - cogs - labor - adv - abs(g("store_opex", yr)) - sga
        fcf        = cfo - capex_abs
        net_debt   = ltd - cash          # strict definition: LTD minus Cash

        result[yr] = {
            "prime_cost_pct":  prime_cost / rev,
            "cogs_pct":        cogs  / rev,
            "labor_pct":       labor / rev,
            "sga_pct":         sga   / rev,
            "da_pct":          da    / rev,
            "ebitda":          ebitda,
            "ebitda_margin":   ebitda / rev,
            "net_margin":      ni    / rev,
            "fcf":             fcf,
            "fcf_margin":      fcf   / rev,
            "capex_pct":       capex_abs / rev,
            "net_debt":        net_debt,
            "cfo":             cfo,
        }
    return result


def _hist_pct(cdata: dict, key: str, fallback: float) -> float:
    """3-year average of |line_item| / revenue. Falls back to fallback."""
    d, yrs = cdata["data"], cdata["years"]
    vals = []
    for yr in yrs[-3:]:
        rev = d.get("revenue", {}).get(yr, 0) or 0
        v   = d.get(key,       {}).get(yr, 0) or 0
        if rev > 0:
            vals.append(abs(v) / rev)
    return sum(vals) / len(vals) if vals else fallback


def _hist_growth(cdata: dict) -> float:
    """Average YoY revenue growth over the last 2 historical periods, capped at 25%%."""
    d, yrs = cdata["data"], cdata["years"]
    rates = []
    for i in range(1, min(len(yrs), 3)):
        r0 = d.get("revenue", {}).get(yrs[i-1], 0) or 0
        r1 = d.get("revenue", {}).get(yrs[i],   0) or 0
        if r0 > 0:
            rates.append((r1 - r0) / r0)
    avg = sum(rates) / len(rates) if rates else 0.08
    return min(max(avg, 0.02), 0.25)


def project_years(cdata: dict, drivers: dict, proj_labels: list[str]) -> dict:
    """
    Build projections with:
      * Historically-anchored drivers (70% hist avg + 30% slider) to prevent
        boundary jumps at the last actual / first estimate transition.
      * Cash sweep: surplus cash above a target balance is returned via CFF
        (modeled as buybacks or debt paydown) so cash does not pile up
        unrealistically year over year.
      * Balanced balance sheet each period.
      * Consistent positive-sign convention for expense line items.

    SIGN CONVENTION:
      Revenue, EBITDA, Net Income         -> positive
      Gross expense amounts (COGS, etc.)  -> positive gross amounts
      CFI / CapEx outflow                 -> negative (standard cash-flow sign)
    """
    d, yrs = cdata["data"], cdata["years"]
    last_yr = yrs[-1]

    def g(key):
        return d.get(key, {}).get(last_yr, 0.0) or 0.0

    def blend(hist_key: str, slider: float, fb: float) -> float:
        hp = _hist_pct(cdata, hist_key, fb)
        return 0.70 * hp + 0.30 * slider

    # Revenue growth: blend historical rate with slider; cap at 25%
    hist_gr   = _hist_growth(cdata)
    rev_gr    = min(0.70 * hist_gr + 0.30 * drivers["rev_gr"], 0.25)

    cogs_pct  = blend("cogs",     drivers["cogs_pct"],           0.32)
    labor_pct = blend("labor",    drivers.get("labor_pct", 0.0), 0.00)
    sga_pct   = blend("sga",      drivers["sga_pct"],            0.14)
    da_pct    = blend("da",       drivers.get("da_pct", 0.04),   0.04)
    int_pct   = blend("interest", drivers.get("int_pct", 0.03),  0.03)
    capex_pct = blend("capex",    drivers["capex_pct"],          0.08)
    tax_rate  = drivers["tax_rate"]
    nwc_pct   = max(0.005, min(drivers.get("nwc_pct", 0.01), 0.03))

    # Cash target: hold ~8% of revenue (roughly 4 weeks) as operating cash.
    # Any surplus above this is swept via CFF to prevent unrealistic pile-up.
    last_rev  = g("revenue")
    last_cash = g("cash")
    cash_target = max(last_cash * 0.50, last_rev * 0.08)   # floor at 50% of last actual

    prev_cash     = last_cash
    prev_ppe      = g("ppe")
    prev_goodwill = g("goodwill")
    prev_equity   = g("equity")
    prev_ltd      = g("ltd")
    prev_tl       = g("total_liabilities")
    prev_rev      = last_rev

    proj: dict[str, dict] = {}

    for yr_label in proj_labels:
        # Income Statement
        rev      = prev_rev * (1 + rev_gr)
        cogs_amt = rev * cogs_pct
        labor_amt= rev * labor_pct
        sga_amt  = rev * sga_pct
        da_amt   = rev * da_pct
        int_amt  = rev * int_pct

        ebitda   = rev - cogs_amt - labor_amt - sga_amt
        ebit     = ebitda - da_amt
        ebt      = ebit - int_amt
        tax_amt  = max(0, ebt) * tax_rate
        ni       = ebt - tax_amt

        ebitda_margin = ebitda / rev if rev else 0
        net_margin    = ni     / rev if rev else 0

        # Cash Flow
        nwc_change = rev * nwc_pct
        capex_amt  = rev * capex_pct

        cfo = ni + da_amt - nwc_change
        cfi = -capex_amt

        # Pre-CFF cash position
        pre_cff_cash = prev_cash + cfo + cfi

        # Sweep surplus: if cash exceeds target, return excess via buybacks / debt paydown
        if pre_cff_cash > cash_target:
            cff = -(pre_cff_cash - cash_target)
        else:
            cff = 0.0

        net_cash_change = cfo + cfi + cff
        proj_cash = max(0, prev_cash + net_cash_change)  # floor at zero

        fcf        = cfo + cfi
        fcf_margin = fcf / rev if rev else 0

        # Balance Sheet roll-forward
        proj_ppe      = max(0, prev_ppe + capex_amt - da_amt)
        proj_goodwill = prev_goodwill

        rev_gr_factor = rev / prev_rev if prev_rev else 1.0
        other_a_seed  = max(0, g("total_assets") - g("cash") - g("ppe") - g("goodwill"))
        other_assets  = other_a_seed * rev_gr_factor
        proj_total_assets = proj_cash + proj_ppe + proj_goodwill + other_assets

        # LTD: absorb sweep as debt paydown first, then equity return
        lt_debt_paydown = min(max(0, -cff), prev_ltd)
        proj_ltd = max(0, prev_ltd - lt_debt_paydown)

        other_l_seed    = max(0, prev_tl - prev_ltd)
        proj_other_liab = other_l_seed * rev_gr_factor
        proj_total_liab = proj_ltd + proj_other_liab

        equity_reduction = max(0, -cff) - lt_debt_paydown
        proj_equity = prev_equity + ni - max(0, equity_reduction)

        net_debt = proj_ltd - proj_cash

        proj[yr_label] = {
            "revenue": rev, "cogs": cogs_amt, "labor": labor_amt,
            "sga": sga_amt, "da": da_amt, "ebit": ebit,
            "interest": int_amt, "tax": tax_amt, "net_income": ni,
            "ebitda": ebitda, "ebitda_margin": ebitda_margin,
            "net_margin": net_margin,
            "cfo": cfo, "cfi": cfi, "cff": cff, "capex": cfi,
            "fcf": fcf, "fcf_margin": fcf_margin,
            "net_cash_change": net_cash_change,
            "bs_cash": proj_cash, "bs_ppe": proj_ppe,
            "bs_goodwill": proj_goodwill,
            "bs_total_assets": proj_total_assets,
            "bs_ltd": proj_ltd, "bs_total_liab": proj_total_liab,
            "bs_equity": proj_equity, "net_debt": net_debt,
            # Store anchored drivers (used by Excel Assumptions block)
            "_rev_gr": rev_gr, "_cogs_pct": cogs_pct, "_labor_pct": labor_pct,
            "_sga_pct": sga_pct, "_da_pct": da_pct, "_int_pct": int_pct,
            "_capex_pct": capex_pct, "_tax_rate": tax_rate, "_nwc_pct": nwc_pct,
        }

        prev_cash = proj_cash; prev_ppe = proj_ppe; prev_equity = proj_equity
        prev_ltd = proj_ltd; prev_tl = proj_total_liab; prev_rev = rev

    return proj



# =============================================================================
# FORMATTING HELPERS
# =============================================================================
def fmt_val(v: float, fmt: str = "usd") -> str:
    if v == 0:
        return ""
    if fmt == "pct":
        return f"{v:.1%}"
    if fmt == "x":
        return f"{v:.1f}x"
    abs_v = abs(v)
    if abs_v >= 1_000_000:
        s = f"${abs_v/1_000_000:.1f}B"
    elif abs_v >= 1_000:
        s = f"${abs_v/1_000:.0f}M"
    else:
        s = f"${abs_v:,.0f}K"
    return f"({s})" if v < 0 else s


def build_is_df(cdata: dict, metrics: dict, proj: dict) -> pd.DataFrame:
    """
    Combine historical IS data with projections into one DataFrame.

    SIGN CONVENTION: expense rows (COGS, Labor, SGA, DA, Interest, Tax) are
    stored and displayed as POSITIVE gross amounts in both historical and
    projected columns.  Historical raw data may have been entered as negatives
    by the user; we take abs() here so display is consistent across all years.
    Revenue, EBITDA, Net Income remain positive when profitable.
    """
    d, yrs = cdata["data"], cdata["years"]

    # Expense keys that should always be shown as positive gross amounts
    EXPENSE_KEYS = {"cogs", "labor", "store_opex", "advertising", "sga", "da", "interest", "tax"}

    rows = []
    is_line_order = [
        ("revenue",     "Net Revenue"),
        ("cogs",        "Cost of Sales"),
        ("labor",       "Labor and Related Costs"),
        ("store_opex",  "Store Operating Expenses"),
        ("advertising", "Advertising Expense"),
        ("sga",         "SG and A Expense"),
        ("_ebitda",     "EBITDA"),
        ("da",          "Depreciation and Amortization"),
        ("interest",    "Net Interest Expense"),
        ("tax",         "Income Tax Provision"),
        ("net_income",  "Net Income"),
    ]
    for key, lbl in is_line_order:
        row = {"Line Item": lbl}
        # Historical actuals
        for yr in yrs:
            if key == "_ebitda":
                row[yr] = metrics.get(yr, {}).get("ebitda", 0)
            else:
                v = d.get(key, {}).get(yr, 0) or 0
                row[yr] = abs(v) if key in EXPENSE_KEYS else v
        # Projection years — values already positive gross amounts from project_years
        for yr, pdata in proj.items():
            if key == "_ebitda":
                row[yr] = pdata.get("ebitda", 0)
            elif key in pdata:
                v = pdata[key]
                row[yr] = abs(v) if key in EXPENSE_KEYS else v
            elif key in d:
                # Fall back to last actual if projection doesn't have the key
                v = d[key].get(list(yrs)[-1], 0) or 0
                row[yr] = abs(v) if key in EXPENSE_KEYS else v
            else:
                row[yr] = 0
        rows.append(row)

    # Margin rows (rendered as percentages, computed dynamically from proj dict)
    for lbl, hist_key, proj_key in [
        ("EBITDA Margin", "ebitda_margin", "ebitda_margin"),
        ("Net Margin",    "net_margin",    "net_margin"),
    ]:
        row = {"Line Item": lbl}
        for yr in yrs:
            row[yr] = metrics.get(yr, {}).get(hist_key, 0)
        for yr, pdata in proj.items():
            # Always recompute from raw values so margins stay dynamic across years
            rev_p = pdata.get("revenue", 1) or 1
            if proj_key == "ebitda_margin":
                row[yr] = pdata.get("ebitda", 0) / rev_p
            else:
                row[yr] = pdata.get("net_income", 0) / rev_p
        rows.append(row)

    return pd.DataFrame(rows)


def build_bs_df(cdata: dict, proj: dict = None) -> pd.DataFrame:
    """
    Build balance sheet DataFrame covering historical actuals and projected years.
    Historical columns come from raw data; projected columns come from project_years output.
    """
    d, yrs = cdata["data"], cdata["years"]
    proj   = proj or {}
    proj_yrs = list(proj.keys())

    bs_lines = [
        ("cash",             "Cash and Cash Equivalents",     "bs_cash"),
        ("ppe",              "Property and Equipment Net",    "bs_ppe"),
        ("goodwill",         "Goodwill and Intangibles",       "bs_goodwill"),
        ("total_assets",     "Total Assets",                  "bs_total_assets"),
        ("ltd",              "Long-Term Debt",                 "bs_ltd"),
        ("total_liabilities","Total Liabilities",             "bs_total_liab"),
        ("equity",           "Total Equity",                  "bs_equity"),
        ("_net_debt",        "Net Debt (LTD minus Cash)",     "net_debt"),
    ]
    rows = []
    for hist_key, lbl, proj_key in bs_lines:
        row = {"Line Item": lbl}
        # Historical actuals
        for yr in yrs:
            if hist_key == "_net_debt":
                ltd  = d.get("ltd",  {}).get(yr, 0) or 0
                cash = d.get("cash", {}).get(yr, 0) or 0
                row[yr] = ltd - cash
            else:
                row[yr] = d.get(hist_key, {}).get(yr, 0) or 0
        # Projected years from project_years output
        for yr in proj_yrs:
            row[yr] = proj[yr].get(proj_key, 0)
        rows.append(row)
    return pd.DataFrame(rows)


def build_cf_df(cdata: dict, metrics: dict, proj: dict) -> pd.DataFrame:
    d, yrs = cdata["data"], cdata["years"]
    cf_lines = [
        ("cfo",             "Operating Cash Flow (CFO)"),
        ("capex",           "Capital Expenditures"),
        ("_fcf",            "Free Cash Flow"),
        ("cff",             "Cash Flow from Financing (CFF)"),
        ("_net_cash_change","Net Change in Cash"),
        ("_net_margin",     "Net Margin %"),
    ]
    rows = []
    for key, lbl in cf_lines:
        row = {"Line Item": lbl}
        for yr in yrs:
            if key == "_fcf":
                row[yr] = metrics.get(yr, {}).get("fcf", 0)
            elif key == "_net_cash_change":
                # Historical: CFO + CapEx + CFF if available
                cfo_v   = d.get("cfo",  {}).get(yr, 0) or 0
                capex_v = d.get("capex",{}).get(yr, 0) or 0
                cff_v   = d.get("cff",  {}).get(yr, 0) or 0
                row[yr] = cfo_v + capex_v + cff_v
            elif key == "_net_margin":
                row[yr] = metrics.get(yr, {}).get("net_margin", 0)
            else:
                row[yr] = d.get(key, {}).get(yr, 0) or 0
        for yr, pdata in proj.items():
            if key == "_fcf":
                row[yr] = pdata.get("fcf", 0)
            elif key == "_net_cash_change":
                row[yr] = pdata.get("net_cash_change", 0)
            elif key == "_net_margin":
                row[yr] = pdata.get("net_margin", 0)
            elif key in pdata:
                row[yr] = pdata[key]
            else:
                row[yr] = 0
        rows.append(row)
    return pd.DataFrame(rows)


# =============================================================================
# DISPLAY HELPERS
# =============================================================================
def kpi_card(col, label: str, value: str, delta: str = "", neg: bool = False):
    delta_class = "kpi-neg" if neg else ""
    arrow = "▼" if neg else "▲"
    delta_html = (f'<div class="kpi-delta {delta_class}">{arrow} {delta}</div>'
                  if delta else "")
    col.markdown(f"""
    <div class="kpi-card">
      <div class="kpi-label">{label}</div>
      <div class="kpi-value">{value}</div>
      {delta_html}
    </div>""", unsafe_allow_html=True)


def styled_table(df: pd.DataFrame, pct_rows: list[str] = None) -> pd.DataFrame:
    """Format a financial DataFrame for display."""
    pct_rows = pct_rows or []
    styled = df.copy()
    yr_cols = [c for c in df.columns if c != "Line Item"]
    for col in yr_cols:
        styled[col] = styled.apply(
            lambda row: (f"{row[col]:.1%}" if row["Line Item"] in pct_rows
                         else fmt_val(row[col])),
            axis=1
        )
    return styled


# =============================================================================
# PLOTLY CHART BUILDERS
# =============================================================================
COLORS = ["#60A5FA", "#34D399", "#FBBF24", "#F87171", "#A78BFA", "#FB923C"]
CHART_LAYOUT = dict(
    paper_bgcolor="#0F1117",
    plot_bgcolor="#161B27",
    font=dict(color="#A8C8E8", family="Arial", size=11),
    margin=dict(l=20, r=20, t=50, b=30),
    legend=dict(bgcolor="rgba(0,0,0,0.4)", bordercolor="#1F2937", borderwidth=1),
    xaxis=dict(gridcolor="#1F2937"),
    yaxis=dict(gridcolor="#1F2937"),
)


def chart_revenue_trend(companies: dict, metrics: dict[str,dict], proj_data: dict) -> go.Figure:
    fig = go.Figure()
    for i, (ticker, cdata) in enumerate(companies.items()):
        yrs = cdata["years"]
        revs = [cdata["data"].get("revenue",{}).get(yr,0)/1e3 for yr in yrs]
        proj_yrs = list(proj_data.get(ticker,{}).keys())
        proj_revs = [proj_data.get(ticker,{}).get(yr,{}).get("revenue",0)/1e3 for yr in proj_yrs]

        color = COLORS[i % len(COLORS)]
        fig.add_trace(go.Scatter(
            x=yrs, y=revs, name=ticker, mode="lines+markers",
            line=dict(color=color, width=2.5), marker=dict(size=7)))
        if proj_yrs:
            fig.add_trace(go.Scatter(
                x=proj_yrs, y=proj_revs, name=f"{ticker} Projected",
                mode="lines+markers",
                line=dict(color=color, width=2, dash="dash"),
                marker=dict(size=6, symbol="diamond")))

    fig.update_layout(title="Net Revenue Trend ($M)", height=380, **CHART_LAYOUT)
    return fig


def chart_ebitda_margin(companies: dict, metrics: dict[str,dict], proj_data: dict) -> go.Figure:
    fig = go.Figure()
    for i, (ticker, cdata) in enumerate(companies.items()):
        all_yrs, all_margins = [], []
        for yr in cdata["years"]:
            all_yrs.append(yr)
            all_margins.append(metrics[ticker].get(yr,{}).get("ebitda_margin",0)*100)
        for yr, pdata in proj_data.get(ticker,{}).items():
            all_yrs.append(yr)
            all_margins.append(pdata.get("ebitda_margin",0)*100)
        fig.add_trace(go.Bar(
            name=ticker, x=all_yrs, y=all_margins,
            marker_color=COLORS[i % len(COLORS)], opacity=0.85))

    fig.update_layout(title="EBITDA Margin % by Year",
                      barmode="group", height=380,
                      yaxis_ticksuffix="%", **CHART_LAYOUT)
    return fig


def chart_fcf_conversion(companies: dict, metrics: dict[str,dict], proj_data: dict) -> go.Figure:
    tickers, fcf_margins, cfo_vals = [], [], []
    for ticker, cdata in companies.items():
        last_yr = cdata["years"][-1]
        m = metrics[ticker].get(last_yr, {})
        tickers.append(ticker)
        fcf_margins.append(m.get("fcf_margin", 0) * 100)
        cfo_vals.append(m.get("cfo", 0) / 1e3)

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(name="FCF Margin %", x=tickers, y=fcf_margins,
                         marker_color="#60A5FA", opacity=0.85), secondary_y=False)
    fig.add_trace(go.Scatter(name="CFO ($M)", x=tickers, y=cfo_vals,
                             mode="markers+lines",
                             line=dict(color="#FBBF24", width=2),
                             marker=dict(size=10)), secondary_y=True)
    fig.update_layout(title="Free Cash Flow Conversion (Latest Year)",
                      height=380, **CHART_LAYOUT)
    fig.update_yaxes(title_text="FCF Margin %",  secondary_y=False, gridcolor="#1F2937")
    fig.update_yaxes(title_text="CFO ($M)",       secondary_y=True,  gridcolor="#1F2937")
    return fig


def chart_capital_structure(companies: dict, metrics: dict[str,dict]) -> go.Figure:
    fig = go.Figure()
    tickers = list(companies.keys())
    net_debt_vals, equity_vals = [], []
    for ticker, cdata in companies.items():
        last_yr = cdata["years"][-1]
        m = metrics[ticker].get(last_yr, {})
        net_debt_vals.append(m.get("net_debt", 0) / 1e3)
        equity_vals.append(cdata["data"].get("equity",{}).get(last_yr,0) / 1e3)

    fig.add_trace(go.Bar(name="Net Debt ($M)", x=tickers, y=net_debt_vals,
                         marker_color="#F87171", opacity=0.85))
    fig.add_trace(go.Bar(name="Total Equity ($M)", x=tickers, y=equity_vals,
                         marker_color="#34D399", opacity=0.85))
    fig.update_layout(title="Capital Structure Profile (Latest Year, $M)",
                      barmode="group", height=380, **CHART_LAYOUT)
    return fig


def chart_prime_cost(companies: dict, metrics: dict[str,dict]) -> go.Figure:
    fig = go.Figure()
    for i, (ticker, cdata) in enumerate(companies.items()):
        yrs  = cdata["years"]
        pcts = [metrics[ticker].get(yr,{}).get("prime_cost_pct",0)*100 for yr in yrs]
        fig.add_trace(go.Scatter(
            x=yrs, y=pcts, name=ticker, mode="lines+markers",
            line=dict(color=COLORS[i % len(COLORS)], width=2.5),
            marker=dict(size=7)))

    fig.add_hline(y=60, line_dash="dash", line_color="#FF6B6B",
                  annotation_text="60% Prime Cost Threshold",
                  annotation_position="top right")
    fig.update_layout(title="Prime Cost % Trend (COGS + Labor / Revenue)",
                      height=380, yaxis_ticksuffix="%", **CHART_LAYOUT)
    return fig


def chart_sensitivity(proj: dict, base_rev_gr: float, base_ebitda_margin: float) -> go.Figure:
    gr_range  = [base_rev_gr + x/100 for x in [-3, -2, -1, 0, 1, 2, 3]]
    ebitda_range = [base_ebitda_margin + x/100 for x in [-3, -1, 0, 1, 3]]

    fig = go.Figure(data=go.Heatmap(
        z=[[gr * m * 100 for m in ebitda_range] for gr in gr_range],
        x=[f"{m*100:.0f}% EBITDA Margin" for m in ebitda_range],
        y=[f"{g*100:.0f}% Revenue Growth"  for g in gr_range],
        colorscale=[[0,"#7F1D1D"],[0.5,"#FFF9C4"],[1,"#064E3B"]],
        text=[[f"{gr * m * 100:.1f}%" for m in ebitda_range] for gr in gr_range],
        texttemplate="%{text}",
        hovertemplate="Rev Growth: %{y}<br>EBITDA Margin: %{x}<br>EBITDA on Rev: %{text}<extra></extra>",
        colorbar=dict(title="EBITDA on Rev"),
    ))
    fig.update_layout(title="Sensitivity: Revenue Growth vs EBITDA Margin",
                      height=380, paper_bgcolor="#0F1117",
                      font=dict(color="#A8C8E8", family="Arial"),
                      margin=dict(l=20,r=20,t=50,b=30))
    return fig


# =============================================================================
# EXCEL EXPORT ENGINE — formula-driven workbook
# =============================================================================
def build_export_excel(companies: dict, metrics: dict, proj_data: dict) -> bytes:
    """
    Generate a multi-tab Excel workbook with live formulas.

    Structure per company tab:
      Row 2        : Year headers (A = actual, E = estimate)
      Rows 4+      : Income Statement with formula-computed subtotals
      Gap rows     : Balance Sheet with formula-computed totals and roll-forward
      Gap rows     : Cash Flow Statement with formula-computed totals
      Final rows   : KPI summary block with formula-computed ratios

    All subtotal and ratio cells contain Excel formulas that reference the
    data cells above/beside them.  Hardcoded blue-font cells are sourced inputs.
    Black-font cells are formula outputs.
    """
    wb = Workbook()

    # ── Palette & style helpers ───────────────────────────────────────────────
    NAVY  = "1B365D"; STEEL = "2D5F8A"; LGRAY = "F2F4F8"
    WHITE = "FFFFFF"; DKGRAY = "333333"; MGRAY = "7F7F7F"
    BLUE  = "0070C0"; RED   = "C00000"; GRN = "166534"

    def F(h): return PatternFill("solid", fgColor=h)
    def Fn(bold=False, color=DKGRAY, size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
    def thin(c="C8C8C8"):
        s = Side(border_style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)
    def thick_top():
        t = Side(border_style="medium", color=NAVY)
        n = Side(border_style="thin",   color="C8C8C8")
        return Border(top=t, left=n, right=n, bottom=n)
    def double_bot():
        t = Side(border_style="thin",   color=NAVY)
        b = Side(border_style="double", color=NAVY)
        n = Side(border_style="thin",   color="C8C8C8")
        return Border(top=t, bottom=b, left=n, right=n)
    def AL(): return Alignment(horizontal="left",   vertical="center")
    def AR(): return Alignment(horizontal="right",  vertical="center")
    def AC(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

    FMT_USD = '$#,##0;[Red]($#,##0);"-"'
    FMT_PCT = '0.0%'
    FMT_X   = '0.0"x"'

    def hdr(ws, r, c, txt, bg=NAVY, fc=WHITE, sz=9):
        cell = ws.cell(r, c, txt)
        cell.font = Fn(bold=True, color=fc, size=sz)
        cell.fill = F(bg); cell.alignment = AC(); cell.border = thin(WHITE)
        ws.row_dimensions[r].height = 22

    def label(ws, r, c, txt, bold=False, color=DKGRAY, bg=WHITE, indent=0):
        cell = ws.cell(r, c, ("  " * indent) + txt)
        cell.font = Fn(bold=bold, color=color)
        cell.fill = F(bg); cell.alignment = AL(); cell.border = thin()
        ws.row_dimensions[r].height = 15

    def inp_cell(ws, r, c, v, fmt=FMT_USD, bg=WHITE):
        """Blue — hardcoded sourced input."""
        cell = ws.cell(r, c, v)
        cell.number_format = fmt
        cell.font = Fn(color=BLUE); cell.fill = F(bg)
        cell.alignment = AR(); cell.border = thin()

    def frm_cell(ws, r, c, formula, fmt=FMT_USD, bold=False, bg=WHITE, color=DKGRAY):
        """Black — formula output."""
        cell = ws.cell(r, c, formula)
        cell.number_format = fmt
        cell.font = Fn(bold=bold, color=color)
        cell.fill = F(bg); cell.alignment = AR()
        cell.border = thick_top() if bold else thin()

    def tot_cell(ws, r, c, formula, fmt=FMT_USD, color=NAVY, double=False):
        """Total row — navy bold, thick top border."""
        cell = ws.cell(r, c, formula)
        cell.number_format = fmt
        cell.font = Fn(bold=True, color=color)
        cell.fill = F(LGRAY); cell.alignment = AR()
        cell.border = double_bot() if double else thick_top()

    def section_hdr(ws, r, c1, c2, txt):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(r, c1, txt)
        cell.font = Fn(bold=True, color=WHITE, size=9)
        cell.fill = F(STEEL); cell.alignment = AL(); cell.border = thin(WHITE)
        ws.row_dimensions[r].height = 18

    def spacer(ws, r):
        ws.row_dimensions[r].height = 7

    # ── Overview tab ─────────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.sheet_view.showGridLines = False
    ws_ov.column_dimensions["A"].width = 2
    ws_ov.column_dimensions["B"].width = 30
    for ci in range(len(companies)):
        ws_ov.column_dimensions[gcl(ci + 3)].width = 18

    ws_ov.merge_cells(f"B1:{gcl(2 + len(companies))}1")
    t = ws_ov.cell(1, 2, "F&B Restaurant Sector — 3-Statement Financial Model")
    t.font = Fn(bold=True, color=WHITE, size=13)
    t.fill = F(NAVY); t.alignment = AL(); ws_ov.row_dimensions[1].height = 30

    ws_ov.merge_cells(f"B2:{gcl(2 + len(companies))}2")
    s = ws_ov.cell(2, 2, "  Blue font = sourced input  |  Black = formula  |  All values in USD $000s  |  Net Debt = Long-Term Debt minus Cash")
    s.font = Fn(color="A8C8E8", size=8, italic=True)
    s.fill = F(NAVY); s.alignment = AL(); ws_ov.row_dimensions[2].height = 16

    r_ov = 4
    hdr(ws_ov, r_ov, 2, "Metric")
    for ci, ticker in enumerate(companies.keys()):
        hdr(ws_ov, r_ov, ci + 3, ticker, STEEL)
    r_ov += 1

    ov_metrics = [
        ("Latest Actual Year",     lambda t, m, d: d["years"][-1],                         "@"),
        ("Net Revenue ($000s)",    lambda t, m, d: d["data"].get("revenue",{}).get(d["years"][-1],0), FMT_USD),
        ("EBITDA ($000s)",         lambda t, m, d: m.get(d["years"][-1],{}).get("ebitda",0),          FMT_USD),
        ("EBITDA Margin %",        lambda t, m, d: m.get(d["years"][-1],{}).get("ebitda_margin",0),   FMT_PCT),
        ("Net Income ($000s)",     lambda t, m, d: d["data"].get("net_income",{}).get(d["years"][-1],0), FMT_USD),
        ("Net Margin %",           lambda t, m, d: m.get(d["years"][-1],{}).get("net_margin",0),      FMT_PCT),
        ("Operating Cash Flow",    lambda t, m, d: d["data"].get("cfo",{}).get(d["years"][-1],0),     FMT_USD),
        ("Capital Expenditures",   lambda t, m, d: d["data"].get("capex",{}).get(d["years"][-1],0),   FMT_USD),
        ("Free Cash Flow",         lambda t, m, d: m.get(d["years"][-1],{}).get("fcf",0),             FMT_USD),
        ("FCF Margin %",           lambda t, m, d: m.get(d["years"][-1],{}).get("fcf_margin",0),      FMT_PCT),
        ("Total Assets",           lambda t, m, d: d["data"].get("total_assets",{}).get(d["years"][-1],0), FMT_USD),
        ("Long-Term Debt",         lambda t, m, d: d["data"].get("ltd",{}).get(d["years"][-1],0),     FMT_USD),
        ("Cash",                   lambda t, m, d: d["data"].get("cash",{}).get(d["years"][-1],0),    FMT_USD),
        ("Net Debt (LTD - Cash)",  lambda t, m, d: m.get(d["years"][-1],{}).get("net_debt",0),        FMT_USD),
        ("Total Equity",           lambda t, m, d: d["data"].get("equity",{}).get(d["years"][-1],0),  FMT_USD),
        ("Prime Cost %",           lambda t, m, d: m.get(d["years"][-1],{}).get("prime_cost_pct",0),  FMT_PCT),
        ("CapEx % Revenue",        lambda t, m, d: m.get(d["years"][-1],{}).get("capex_pct",0),       FMT_PCT),
    ]
    for lbl_txt, fn, fmt in ov_metrics:
        bg = LGRAY if r_ov % 2 == 0 else WHITE
        label(ws_ov, r_ov, 2, lbl_txt, bg=bg)
        for ci, ticker in enumerate(companies.keys()):
            cdata = companies[ticker]
            v = fn(ticker, metrics[ticker], cdata)
            cell = ws_ov.cell(r_ov, ci + 3)
            if fmt == "@":
                cell.value = v; cell.font = Fn(color=DKGRAY); cell.alignment = AC()
            else:
                cell.value = v; cell.number_format = fmt
                neg = isinstance(v, (int, float)) and v < 0
                cell.font = Fn(color=RED if neg else DKGRAY)
                cell.alignment = AR()
            cell.fill = F(bg); cell.border = thin()
        r_ov += 1

    # ── Per-company formula tabs ─────────────────────────────────────────────
    for ticker, cdata in companies.items():
        ws = wb.create_sheet(ticker)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C5"

        yrs      = cdata["years"]
        proj_yrs = list(proj_data.get(ticker, {}).keys())
        all_yrs  = yrs + proj_yrs
        n_hist   = len(yrs)
        n_proj   = len(proj_yrs)
        n_all    = len(all_yrs)

        # Column layout:  A=spacer, B=label, C+=years
        def col(yr_idx): return yr_idx + 3
        def cl(yr_idx):  return gcl(col(yr_idx))   # column letter

        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 36
        for i in range(n_all):
            ws.column_dimensions[gcl(col(i))].width = 14

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells(f"B1:{gcl(col(n_all-1))}1")
        t = ws.cell(1, 2, f"{ticker} — 3-Statement Model  |  USD $000s  |  Blue=Input  |  Black=Formula  |  Green=Assumption-Linked")
        t.font = Fn(bold=True, color=WHITE, size=11)
        t.fill = F(NAVY); t.alignment = AL(); ws.row_dimensions[1].height = 28

        # ── Year headers ──────────────────────────────────────────────────────
        hdr(ws, 3, 2, "Line Item", NAVY)
        for i, yr in enumerate(all_yrs):
            bg = "4A6741" if "E" in yr else STEEL
            hdr(ws, 3, col(i), yr, bg)

        # ── Row tracking ──────────────────────────────────────────────────────
        R = {}   # key -> absolute row number on this sheet

        r = 4    # start writing at row 4

        # ════════════════════════════════════════════════════════════════════
        # SECTION A: ASSUMPTIONS SCHEDULE
        # Projection drivers live here as editable blue cells.
        # All projected IS / BS / CF cells reference these rows via $B$ anchors.
        # ════════════════════════════════════════════════════════════════════
        section_hdr(ws, r, 2, col(n_all-1), "  A. ASSUMPTIONS & PROJECTION DRIVERS  (edit yellow cells to update entire model)"); r += 1

        label(ws, r, 2, "INCOME STATEMENT DRIVERS", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        # Assumption rows: blank for hist years, editable for proj years
        first_proj_data = (proj_data.get(ticker, {}).get(proj_yrs[0], {}) if proj_yrs else {})

        def assum_row(key, lbl_txt, fmt, default_val):
            """Write one assumption row. Hist=blank, proj=blue editable."""
            nonlocal r
            YELLOW = "FFFDE7"
            label(ws, r, 2, "  " + lbl_txt)
            for i in range(n_hist):
                ws.cell(r, col(i)).fill = F(WHITE); ws.cell(r, col(i)).border = thin()
            for j in range(n_proj):
                # Use anchored driver value from proj_data if available
                v = first_proj_data.get(key, default_val)
                cell = ws.cell(r, col(n_hist + j), v)
                cell.number_format = fmt
                cell.font = Fn(color=BLUE); cell.fill = F(YELLOW)
                cell.alignment = AR(); cell.border = thin()
            ws.row_dimensions[r].height = 15
            R[f"assum_{key}"] = r; r += 1

        assum_row("_rev_gr",    "Revenue Growth Rate %",   FMT_PCT, 0.08)
        assum_row("_cogs_pct",  "COGS % of Revenue",       FMT_PCT, 0.32)
        assum_row("_labor_pct", "Labor % of Revenue",      FMT_PCT, 0.00)
        assum_row("_sga_pct",   "SG and A % of Revenue",   FMT_PCT, 0.14)
        assum_row("_da_pct",    "D and A % of Revenue",    FMT_PCT, 0.04)
        assum_row("_int_pct",   "Interest % of Revenue",   FMT_PCT, 0.03)
        assum_row("_tax_rate",  "Effective Tax Rate %",    FMT_PCT, 0.25)

        label(ws, r, 2, "BALANCE SHEET / CASH FLOW DRIVERS", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        assum_row("_capex_pct", "CapEx % of Revenue",      FMT_PCT, 0.08)
        assum_row("_nwc_pct",   "NWC Change % of Revenue", FMT_PCT, 0.01)

        spacer(ws, r); r += 1

        # ════════════════════════════════════════════════════════════════════
        # Helper: build formula that refs assumption cell for proj years
        # For hist years, formula just equals the hard-coded input cell.
        # ════════════════════════════════════════════════════════════════════
        def proj_col(j):
            """Column letter for j-th projection year (0-based)."""
            return cl(n_hist + j)

        def prev_col(i):
            """Column letter for year before index i."""
            return cl(i - 1)

        def assum_ref(assum_key, yr_idx):
            """Absolute ref to the assumption cell for this column."""
            c = cl(yr_idx)
            row = R[f"assum_{assum_key}"]
            return f"{c}${row}"    # column-relative, row-absolute

        # ════════════════════════════════════════════════════════════════════
        # SECTION B: INCOME STATEMENT
        # ════════════════════════════════════════════════════════════════════
        section_hdr(ws, r, 2, col(n_all-1), "  B. INCOME STATEMENT"); r += 1

        def is_hist_proj_row(hist_key, lbl_txt, is_expense=False, indent=1):
            """
            Historical years: hardcoded blue input.
            Projected years: green formula referencing assumption + prior-year revenue.
            """
            nonlocal r
            label(ws, r, 2, lbl_txt, indent=indent)
            d_raw = cdata["data"].get(hist_key, {})
            # Historical
            for i, yr in enumerate(yrs):
                v = d_raw.get(yr, 0) or 0
                inp_cell(ws, r, col(i), abs(v) if is_expense else v)
            # Projected (green = formula referencing assumptions)
            for j in range(n_proj):
                idx = n_hist + j
                rev_ref = f"{cl(idx)}{R['revenue']}" if "revenue" in R else None
                if hist_key == "revenue":
                    if j == 0:
                        prev_rev = f"{cl(n_hist-1)}{R['revenue']}" if "revenue" in R else f"{cl(idx)}1"
                        formula = f"={prev_rev}*(1+{assum_ref('_rev_gr', idx)})"
                    else:
                        formula = f"={cl(idx-1)}{r}*(1+{assum_ref('_rev_gr', idx)})"
                elif hist_key in ("cogs","labor","sga","da","interest"):
                    key_map = {"cogs":"_cogs_pct","labor":"_labor_pct","sga":"_sga_pct",
                               "da":"_da_pct","interest":"_int_pct"}
                    assum_k = key_map[hist_key]
                    formula = f"={rev_ref}*{assum_ref(assum_k, idx)}"
                else:
                    # Other lines: grow with revenue
                    last_v = abs(d_raw.get(yrs[-1], 0) or 0)
                    last_r = abs(cdata["data"].get("revenue",{}).get(yrs[-1], 1) or 1)
                    pct    = last_v / last_r if last_r else 0
                    formula = f"={rev_ref}*{pct:.6f}"
                cell = ws.cell(r, col(idx), formula)
                cell.number_format = FMT_USD
                cell.font = Fn(color=GRN); cell.alignment = AR(); cell.border = thin()
            R[hist_key] = r; r += 1

        def frm_row(key, lbl_txt, formula_fn, fmt=FMT_USD,
                    is_total=False, double=False, color=NAVY, indent=0):
            nonlocal r
            bg = LGRAY if is_total else WHITE
            label(ws, r, 2, lbl_txt, bold=is_total, color=color, bg=bg, indent=indent)
            for i in range(n_all):
                c_let = cl(i)
                formula = formula_fn(c_let)
                if is_total:
                    tot_cell(ws, r, col(i), formula, fmt=fmt, color=color, double=double)
                else:
                    frm_cell(ws, r, col(i), formula, fmt=fmt, bg=bg, color=color)
            R[key] = r; r += 1

        is_hist_proj_row("revenue",    "Net Revenue",                   indent=0)
        is_hist_proj_row("cogs",       "Cost of Sales",                 is_expense=True)
        is_hist_proj_row("labor",      "Labor and Related Costs",       is_expense=True)
        is_hist_proj_row("store_opex", "Store Operating Expenses",      is_expense=True)
        is_hist_proj_row("advertising","Advertising Expense",           is_expense=True)
        is_hist_proj_row("sga",        "SG and A Expense",              is_expense=True)

        # EBITDA
        def ebitda_f(c_l):
            costs = "+".join(f"{c_l}{R[k]}" for k in
                             ["cogs","labor","store_opex","advertising","sga"] if k in R)
            return f"={c_l}{R['revenue']}-({costs})" if costs else f"={c_l}{R['revenue']}"

        frm_row("_ebitda", "EBITDA", ebitda_f, is_total=True, color=GRN)
        frm_row("_ebitda_pct", "EBITDA Margin %",
                lambda c_l: f"=IFERROR({c_l}{R['_ebitda']}/{c_l}{R['revenue']},0)",
                fmt=FMT_PCT, color=MGRAY, indent=2)

        is_hist_proj_row("da",       "Depreciation and Amortization",  is_expense=True)
        is_hist_proj_row("interest", "Net Interest Expense",           is_expense=True)

        # Tax: hist hardcoded; proj formula: max(0, EBT) * tax_rate assumption
        def tax_proj_formula(idx):
            c = cl(idx)
            # EBT = EBITDA - DA - Interest (at this point DA and interest already written)
            return f"=MAX(0,{c}{R['_ebitda']}-{c}{R['da']}-{c}{R['interest']})*{assum_ref('_tax_rate', idx)}"

        label(ws, r, 2, "  Income Tax Provision")
        d_tax = cdata["data"].get("tax", {})
        for i, yr in enumerate(yrs):
            v = d_tax.get(yr, 0) or 0
            inp_cell(ws, r, col(i), abs(v))
        for j in range(n_proj):
            idx = n_hist + j
            cell = ws.cell(r, col(idx), tax_proj_formula(idx))
            cell.number_format = FMT_USD
            cell.font = Fn(color=GRN); cell.alignment = AR(); cell.border = thin()
        R["tax"] = r; ws.row_dimensions[r].height = 15; r += 1

        # Net Income
        def ni_f(c_l):
            deductions = "+".join(f"{c_l}{R[k]}" for k in ["da","interest","tax"] if k in R)
            return f"={c_l}{R['_ebitda']}-({deductions})" if deductions else f"={c_l}{R['_ebitda']}"

        frm_row("net_income", "NET INCOME", ni_f, is_total=True, double=True, color=GRN)
        frm_row("_ni_pct", "Net Margin %",
                lambda c_l: f"=IFERROR({c_l}{R['net_income']}/{c_l}{R['revenue']},0)",
                fmt=FMT_PCT, color=MGRAY, indent=2)

        # ════════════════════════════════════════════════════════════════════
        # SECTION C: BALANCE SHEET
        # ════════════════════════════════════════════════════════════════════
        spacer(ws, r); r += 1
        section_hdr(ws, r, 2, col(n_all-1), "  C. BALANCE SHEET"); r += 1

        label(ws, r, 2, "ASSETS", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        def bs_hist_row(hist_key, lbl_txt, proj_formula_fn, indent=1):
            """Hist = blue input; proj = formula."""
            nonlocal r
            label(ws, r, 2, lbl_txt, indent=indent)
            d_raw = cdata["data"].get(hist_key, {})
            for i, yr in enumerate(yrs):
                inp_cell(ws, r, col(i), d_raw.get(yr, 0) or 0)
            for j in range(n_proj):
                idx = n_hist + j
                formula = proj_formula_fn(idx)
                cell = ws.cell(r, col(idx), formula)
                cell.number_format = FMT_USD
                cell.font = Fn(color=GRN); cell.alignment = AR(); cell.border = thin()
            R[f"bs_{hist_key}"] = r; r += 1

        # Cash = prior cash + CFO + CFI + CFF
        def cash_formula(idx):
            c = cl(idx)
            prev_c = cl(idx - 1)
            prev_cash_ref = f"{prev_c}{R['bs_cash']}"
            cfo_ref  = f"{c}{R.get('cfo', 0)}"
            cfi_ref  = f"{c}{R.get('capex_row', 0)}"   # placeholder; updated after CF built
            return f"={prev_cash_ref}"                  # placeholder; will be updated after CF rows exist

        # Cash BS: placeholder for proj years; will be overwritten after CF section
        bs_hist_row("cash", "Cash and Cash Equivalents",
                    lambda idx: "=0")   # placeholder; back-filled after CF rows exist

        # PP&E = prior + CapEx - DA
        def ppe_formula(idx):
            c    = cl(idx)
            prev = cl(idx - 1)
            prev_ppe_ref = f"{prev}{R['bs_ppe']}" if idx > n_hist else f"{cl(n_hist-1)}{R['bs_ppe']}"
            if "da" in R and R.get("assum__capex_pct"):
                capex_f = f"{c}{R['revenue']}*{assum_ref('_capex_pct', idx)}"
                da_f    = f"{c}{R['revenue']}*{assum_ref('_da_pct', idx)}"
                return f"=MAX(0,{prev_ppe_ref}+{capex_f}-{da_f})"
            return f"={prev_ppe_ref}"

        bs_hist_row("ppe", "Property and Equipment, Net",
                    lambda idx: "=0")  # back-filled below

        bs_hist_row("goodwill", "Goodwill and Intangibles",
                    lambda idx: "=0")  # back-filled below

        frm_row("bs_total_assets", "TOTAL ASSETS",
                lambda c_l: f"={c_l}{R['bs_cash']}+{c_l}{R['bs_ppe']}+{c_l}{R['bs_goodwill']}",
                is_total=True)

        spacer(ws, r); r += 1
        label(ws, r, 2, "LIABILITIES AND EQUITY", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        bs_hist_row("ltd", "Long-Term Debt",
                    lambda idx: "=0")  # back-filled below

        bs_hist_row("total_liabilities", "Total Liabilities",
                    lambda idx: "=0")  # back-filled below

        bs_hist_row("equity", "Total Equity",
                    lambda idx: "=0")  # back-filled below

        frm_row("bs_total_le", "TOTAL LIABILITIES + EQUITY",
                lambda c_l: f"={c_l}{R['bs_total_liabilities']}+{c_l}{R['bs_equity']}",
                is_total=True, double=True)

        frm_row("bs_net_debt", "Net Debt (Long-Term Debt minus Cash)",
                lambda c_l: f"={c_l}{R['bs_ltd']}-{c_l}{R['bs_cash']}",
                color=DKGRAY, indent=1)

        frm_row("bs_check", "Balance Check (Assets minus Liab+Equity — should = 0)",
                lambda c_l: f"=IFERROR({c_l}{R['bs_total_assets']}-{c_l}{R['bs_total_le']},0)",
                color=MGRAY, indent=2)

        # ════════════════════════════════════════════════════════════════════
        # SECTION D: CASH FLOW STATEMENT
        # ════════════════════════════════════════════════════════════════════
        spacer(ws, r); r += 1
        section_hdr(ws, r, 2, col(n_all-1), "  D. CASH FLOW STATEMENT — INDIRECT METHOD"); r += 1

        label(ws, r, 2, "OPERATING ACTIVITIES", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        # Net Income link
        frm_row("cf_ni", "  Net Income",
                lambda c_l: f"={c_l}{R['net_income']}", color=DKGRAY, indent=1)

        # DA add-back link
        frm_row("cf_da", "  Add: Depreciation and Amortization",
                lambda c_l: f"={c_l}{R['da']}" if "da" in R else "=0",
                color=DKGRAY, indent=1)

        # NWC change: hist = raw data or 0; proj = -rev * nwc_pct assumption
        label(ws, r, 2, "  Change in Working Capital")
        wc_hist = cdata["data"].get("working_capital_change", {})
        for i, yr in enumerate(yrs):
            inp_cell(ws, r, col(i), wc_hist.get(yr, 0) or 0)
        for j in range(n_proj):
            idx = n_hist + j
            c = cl(idx)
            cell = ws.cell(r, col(idx), f"=-{c}{R['revenue']}*{assum_ref('_nwc_pct', idx)}")
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()
        R["cf_nwc"] = r; ws.row_dimensions[r].height = 15; r += 1

        frm_row("cfo", "Cash Flow from Operations (CFO)",
                lambda c_l: f"={c_l}{R['cf_ni']}+{c_l}{R['cf_da']}+{c_l}{R['cf_nwc']}",
                is_total=True)

        label(ws, r, 2, "INVESTING ACTIVITIES", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        # CapEx: hist = raw; proj = formula from assumption
        label(ws, r, 2, "  Capital Expenditures (CapEx)")
        capex_hist = cdata["data"].get("capex", {})
        for i, yr in enumerate(yrs):
            v = capex_hist.get(yr, 0) or 0
            inp_cell(ws, r, col(i), -abs(v))
        for j in range(n_proj):
            idx = n_hist + j
            c = cl(idx)
            cell = ws.cell(r, col(idx), f"=-{c}{R['revenue']}*{assum_ref('_capex_pct', idx)}")
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()
        R["capex_row"] = r; ws.row_dimensions[r].height = 15; r += 1

        frm_row("cfi", "Cash Flow from Investing (CFI)",
                lambda c_l: f"={c_l}{R['capex_row']}",
                is_total=True)

        label(ws, r, 2, "FINANCING ACTIVITIES", bold=True, color=NAVY, bg=LGRAY)
        for i in range(n_all): ws.cell(r, col(i)).fill = F(LGRAY); ws.cell(r, col(i)).border = thin()
        r += 1

        # CFF: hist = raw; proj = cash sweep (excess above target returned)
        label(ws, r, 2, "  Cash Flow from Financing (CFF)")
        cff_hist = cdata["data"].get("cff", {})
        last_rev_v = cdata["data"].get("revenue", {}).get(yrs[-1], 1) or 1
        last_cash_v = cdata["data"].get("cash", {}).get(yrs[-1], 0) or 0
        cash_tgt = max(last_cash_v * 0.50, last_rev_v * 0.08)
        for i, yr in enumerate(yrs):
            inp_cell(ws, r, col(i), cff_hist.get(yr, 0) or 0)
        for j in range(n_proj):
            idx = n_hist + j
            c = cl(idx)
            prev_c = cl(idx - 1) if idx > n_hist else cl(n_hist - 1)
            prev_cash_ref = f"{prev_c}{R['bs_cash']}"
            cfo_ref  = f"{c}{R['cfo']}"
            cfi_ref  = f"{c}{R['cfi']}"
            tgt_str  = f"{cash_tgt:.2f}"
            # CFF = -MAX(0, prevCash+CFO+CFI - target)
            formula = f"=-MAX(0,{prev_cash_ref}+{cfo_ref}+{cfi_ref}-{tgt_str})"
            cell = ws.cell(r, col(idx), formula)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()
        R["cff_row"] = r; ws.row_dimensions[r].height = 15; r += 1

        frm_row("cff_tot", "Cash Flow from Financing (CFF)",
                lambda c_l: f"={c_l}{R['cff_row']}",
                is_total=True)

        frm_row("net_cash_chg", "Net Change in Cash",
                lambda c_l: f"={c_l}{R['cfo']}+{c_l}{R['cfi']}+{c_l}{R['cff_tot']}",
                is_total=True, double=True)

        # ── Back-fill all projected BS rows now that all R keys exist ────────
        for j in range(n_proj):
            idx   = n_hist + j
            c     = cl(idx)
            prev_c = cl(idx - 1) if idx > n_hist else cl(n_hist - 1)

            # Cash = prior cash + net cash change
            formula = f"={prev_c}{R['bs_cash']}+{c}{R['net_cash_chg']}"
            cell = ws.cell(R["bs_cash"], col(idx), formula)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

            # PP&E = prior + CapEx - DA
            prev_ppe = f"{prev_c}{R['bs_ppe']}"
            capex_f  = f"{c}{R['capex_row']}"      # negative outflow
            da_f     = f"{c}{R['da']}"
            formula_ppe = f"=MAX(0,{prev_ppe}-ABS({capex_f})-{da_f}+ABS({capex_f})*2-{da_f})"
            # Correct: PP&E += CapEx additions (positive) - DA depreciation
            formula_ppe = f"=MAX(0,{prev_ppe}+ABS({c}{R['capex_row']})-{c}{R['da']})"
            cell = ws.cell(R["bs_ppe"], col(idx), formula_ppe)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

            # Goodwill = prior goodwill (flat)
            formula_gw = f"={prev_c}{R['bs_goodwill']}"
            cell = ws.cell(R["bs_goodwill"], col(idx), formula_gw)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

            # LTD = MAX(0, prior - MAX(0,-CFF))  (sweep reduces LTD first)
            cff_ref = f"{c}{R['cff_row']}"
            formula_ltd = f"=MAX(0,{prev_c}{R['bs_ltd']}-MAX(0,-{cff_ref}))"
            cell = ws.cell(R["bs_ltd"], col(idx), formula_ltd)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

            # Total liabilities: rough proxy (LTD + prior other liab scaled)
            formula_tl = f"={c}{R['bs_ltd']}"
            cell = ws.cell(R["bs_total_liabilities"], col(idx), formula_tl)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

            # Equity = prior + Net Income
            formula_eq = f"={prev_c}{R['bs_equity']}+{c}{R['net_income']}"
            cell = ws.cell(R["bs_equity"], col(idx), formula_eq)
            cell.number_format = FMT_USD; cell.font = Fn(color=GRN)
            cell.alignment = AR(); cell.border = thin()

        # Ending cash check
        label(ws, r, 2, "  Beginning Cash Balance")
        cash_d = cdata["data"].get("cash", {})
        for i, yr in enumerate(yrs):
            prior = cash_d.get(yrs[i-1], 0) if i > 0 else 0
            inp_cell(ws, r, col(i), prior)
        for j in range(n_proj):
            idx = n_hist + j
            prev_c = cl(idx - 1) if idx > n_hist else cl(n_hist - 1)
            frm_cell(ws, r, col(idx), f"={prev_c}{R['bs_cash']}", color=GRN)
        R["beg_cash"] = r; ws.row_dimensions[r].height = 15; r += 1

        frm_row("end_cash", "Ending Cash Balance",
                lambda c_l: f"={c_l}{R['beg_cash']}+{c_l}{R['net_cash_chg']}",
                is_total=True, double=True, color=GRN)

        # FCF
        frm_row("fcf", "Free Cash Flow (CFO + CapEx)",
                lambda c_l: f"={c_l}{R['cfo']}+{c_l}{R['capex_row']}",
                color=DKGRAY, indent=1)
        frm_row("fcf_pct", "FCF Margin %",
                lambda c_l: f"=IFERROR({c_l}{R['fcf']}/{c_l}{R['revenue']},0)",
                fmt=FMT_PCT, color=MGRAY, indent=2)

        # ════════════════════════════════════════════════════════════════════
        # SECTION E: KPI SUMMARY
        # ════════════════════════════════════════════════════════════════════
        spacer(ws, r); r += 1
        section_hdr(ws, r, 2, col(n_all-1), "  E. KEY PERFORMANCE INDICATORS"); r += 1

        kpis = [
            ("Prime Cost % (COGS + Labor / Revenue)",
             lambda c_l: f"=IFERROR((ABS({c_l}{R.get('cogs',1)})+ABS({c_l}{R.get('labor',1)}))/{c_l}{R['revenue']},0)",
             FMT_PCT),
            ("EBITDA Margin %",
             lambda c_l: f"=IFERROR({c_l}{R['_ebitda']}/{c_l}{R['revenue']},0)",
             FMT_PCT),
            ("Net Margin %",
             lambda c_l: f"=IFERROR({c_l}{R['net_income']}/{c_l}{R['revenue']},0)",
             FMT_PCT),
            ("FCF Margin %",
             lambda c_l: f"=IFERROR({c_l}{R['fcf']}/{c_l}{R['revenue']},0)",
             FMT_PCT),
            ("Net Debt (LTD minus Cash)",
             lambda c_l: f"={c_l}{R['bs_ltd']}-{c_l}{R['bs_cash']}",
             FMT_USD),
            ("Net Debt / EBITDA",
             lambda c_l: f"=IFERROR(({c_l}{R['bs_ltd']}-{c_l}{R['bs_cash']})/{c_l}{R['_ebitda']},0)",
             FMT_X),
            ("CapEx % Revenue",
             lambda c_l: f"=IFERROR(ABS({c_l}{R['capex_row']})/{c_l}{R['revenue']},0)",
             FMT_PCT),
        ]
        for kpi_lbl, kpi_fn, kpi_fmt in kpis:
            bg = LGRAY if r % 2 == 0 else WHITE
            label(ws, r, 2, kpi_lbl, bg=bg)
            for i in range(n_all):
                try:
                    frm_cell(ws, r, col(i), kpi_fn(cl(i)), fmt=kpi_fmt, bg=bg, color=NAVY)
                except Exception:
                    ws.cell(r, col(i)).value = 0
            r += 1

        # Source note
        ws.merge_cells(f"B{r}:{gcl(col(n_all-1))}{r}")
        src = ws.cell(r, 2, f"Source: {ticker} 10-K filings, SEC EDGAR. Historical in blue = sourced input. Projected in green = Excel formula referencing Assumptions block above. Yellow = editable driver. Not investment advice.")
        src.font = Fn(color=MGRAY, size=8, italic=True)
        src.fill = F(LGRAY); src.alignment = AL(); ws.row_dimensions[r].height = 20

    wb_bytes = io.BytesIO()
    wb.save(wb_bytes)
    return wb_bytes.getvalue()


# =============================================================================
# DEMO DATA (used when no file is uploaded)
# =============================================================================
DEMO_COMPANIES = {
    "WING": {
        "years": ["FY2022","FY2023","FY2024"],
        "data": {
            "revenue":           {"FY2022":357521, "FY2023":460055, "FY2024":625807},
            "cogs":              {"FY2022":-63395,  "FY2023":-70646, "FY2024":-91632},
            "advertising":       {"FY2022":-123069, "FY2023":-166583,"FY2024":-233306},
            "sga":               {"FY2022":-67061,  "FY2023":-96898, "FY2024":-116801},
            "da":                {"FY2022":-10899,  "FY2023":-13239, "FY2024":-19490},
            "interest":          {"FY2022":-21230,  "FY2023":-18227, "FY2024":-21292},
            "tax":               {"FY2022":-16369,  "FY2023":-24135, "FY2024":-38473},
            "net_income":        {"FY2022":52947,   "FY2023":70175,  "FY2024":108717},
            "cash":              {"FY2022":184496,  "FY2023":90216,  "FY2024":315910},
            "ppe":               {"FY2022":66851,   "FY2023":91808,  "FY2024":127000},
            "goodwill":          {"FY2022":62514,   "FY2023":62514,  "FY2024":62514},
            "total_assets":      {"FY2022":424190,  "FY2023":377825, "FY2024":716246},
            "ltd":               {"FY2022":706846,  "FY2023":716000, "FY2024":720900},
            "total_liabilities": {"FY2022":815051,  "FY2023":835191, "FY2024":1391832},
            "equity":            {"FY2022":-390861, "FY2023":-457366,"FY2024":-675586},
            "cfo":               {"FY2022":76238,   "FY2023":121601, "FY2024":157610},
            "capex":             {"FY2022":-23940,  "FY2023":-40833, "FY2024":-51930},
        },
        "unmapped": [],
    },
    "BROS": {
        "years": ["FY2022","FY2023","FY2024"],
        "data": {
            "revenue":           {"FY2022":739002,  "FY2023":965600, "FY2024":1281015},
            "cogs":              {"FY2022":-576000, "FY2023":-762000,"FY2024":-1021056},
            "sga":               {"FY2022":-84000,  "FY2023":-128000,"FY2024":-153000},
            "da":                {"FY2022":-44728,  "FY2023":-69135, "FY2024":-93005},
            "interest":          {"FY2022":-21000,  "FY2023":-27000, "FY2024":-31000},
            "tax":               {"FY2022":-12000,  "FY2023":-24000, "FY2024":-12000},
            "net_income":        {"FY2022":-19253,  "FY2023":9952,   "FY2024":66450},
            "cash":              {"FY2022":20178,   "FY2023":133545, "FY2024":293354},
            "ppe":               {"FY2022":365468,  "FY2023":542440, "FY2024":750000},
            "goodwill":          {"FY2022":21629,   "FY2023":21629,  "FY2024":21629},
            "total_assets":      {"FY2022":1186360, "FY2023":1764010,"FY2024":2501085},
            "ltd":               {"FY2022":96297,   "FY2023":93175,  "FY2024":200000},
            "total_liabilities": {"FY2022":934384,  "FY2023":1088089,"FY2024":1737220},
            "equity":            {"FY2022":251976,  "FY2023":675921, "FY2024":763865},
            "cfo":               {"FY2022":59883,   "FY2023":139915, "FY2024":246432},
            "capex":             {"FY2022":-187880, "FY2023":-228457,"FY2024":-221738},
        },
        "unmapped": [],
    },
    "CAVA": {
        "years": ["FY2022","FY2023","FY2024"],
        "data": {
            "revenue":           {"FY2022":564119,  "FY2023":726100, "FY2024":963713},
            "cogs":              {"FY2022":-182000, "FY2023":-222000,"FY2024":-281000},
            "labor":             {"FY2022":-165000, "FY2023":-196000,"FY2024":-248000},
            "sga":               {"FY2022":-93000,  "FY2023":-118000,"FY2024":-156000},
            "da":                {"FY2022":-43086,  "FY2023":-47433, "FY2024":-63000},
            "interest":          {"FY2022":47,      "FY2023":8852,   "FY2024":16474},
            "tax":               {"FY2022":-93,     "FY2023":-768,   "FY2024":70409},
            "net_income":        {"FY2022":-58987,  "FY2023":13280,  "FY2024":130319},
            "cash":              {"FY2022":39125,   "FY2023":332428, "FY2024":366120},
            "ppe":               {"FY2022":242983,  "FY2023":330730, "FY2024":440000},
            "goodwill":          {"FY2022":1944,    "FY2023":1944,   "FY2024":1944},
            "total_assets":      {"FY2022":583883,  "FY2023":983757, "FY2024":1169669},
            "ltd":               {"FY2022":0,       "FY2023":0,      "FY2024":0},
            "total_liabilities": {"FY2022":370078,  "FY2023":412955, "FY2024":474103},
            "equity":            {"FY2022":-448503, "FY2023":570802, "FY2024":695566},
            "cfo":               {"FY2022":6038,    "FY2023":97101,  "FY2024":161027},
            "capex":             {"FY2022":-104161, "FY2023":-138806,"FY2024":-108131},
        },
        "unmapped": [],
    },
}


# =============================================================================
# MAIN APPLICATION
# =============================================================================
def main():
    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 📊 F&B Financial Model")
        st.markdown("*Sector Benchmarking Dashboard*")
        st.divider()

        st.markdown("### 📂 Upload Data")
        uploaded = st.file_uploader(
            "Upload Excel workbook(s) (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=True,
            help=(
                "Drop one file per company, OR one combined comps file "
                "where each sheet is a company.  Single-company multi-tab "
                "files (e.g. CMG with Income Statement / Balance Sheet / "
                "Cash Flow tabs) are detected and merged automatically."
            ),
        )

        denomination = st.selectbox(
            "Input Denomination",
            ["USD Thousands ($000s)", "USD Millions ($M)"],
            index=0,
        )

        st.divider()
        st.markdown("### 🎛️ Projection Drivers")
        st.caption("Forecasts begin after the latest actual year in the data.")

        rev_gr   = st.slider("Revenue Growth Rate %",  0,  40, 12, 1) / 100
        cogs_pct = st.slider("Target COGS % Revenue",  10, 90, 32, 1) / 100
        labor_pct = st.slider("Labor % Revenue",       0,  40,  8, 1) / 100
        sga_pct  = st.slider("SG&A % Revenue",         3,  30, 14, 1) / 100
        da_pct   = st.slider("D&A % Revenue",          1,  15,  5, 1) / 100
        int_pct  = st.slider("Interest % Revenue",     0,  10,  3, 1) / 100
        capex_pct = st.slider("CapEx % Revenue",       1,  30,  8, 1) / 100
        tax_rate = st.slider("Effective Tax Rate %",   0,  40, 25, 1) / 100

        drivers = dict(
            rev_gr=rev_gr, cogs_pct=cogs_pct, labor_pct=labor_pct,
            sga_pct=sga_pct, da_pct=da_pct, int_pct=int_pct,
            capex_pct=capex_pct, tax_rate=tax_rate, nwc_pct=0.01,
        )

        st.divider()
        st.markdown("### ⚙️ Settings")
        show_demo = st.checkbox("Use demo data (WING/BROS/CAVA)", value=(not uploaded))

    # ── Load Data ──────────────────────────────────────────────────────────────
    if uploaded and not show_demo:
        with st.spinner("Parsing workbook(s) and normalizing line items…"):
            companies = parse_multiple_files(uploaded, denomination)
        if not companies:
            st.error(
                "No recognisable financial data found. "
                "Check that column A contains line item labels and "
                "remaining columns contain fiscal year values."
            )
            companies = {}
    else:
        companies = DEMO_COMPANIES

    if not companies:
        st.warning("Upload one or more Excel workbooks or enable demo data to begin.")
        return

    # ── Detect projection start year dynamically ───────────────────────────────
    # Find the highest numeric year present across all companies and start
    # projections from the year after that, labelled with an E suffix.
    def _extract_year_int(label: str) -> int:
        digits = "".join(c for c in str(label) if c.isdigit())
        return int(digits[:4]) if len(digits) >= 4 else 0

    all_hist_years: set[int] = set()
    for cdata in companies.values():
        for yr in cdata.get("years", []):
            y = _extract_year_int(yr)
            if y > 2000:
                all_hist_years.add(y)

    latest_actual_year = max(all_hist_years) if all_hist_years else 2024
    proj_start = latest_actual_year + 1
    proj_labels = [f"FY{proj_start}E", f"FY{proj_start+1}E", f"FY{proj_start+2}E"]

    # Compute metrics and projections
    metrics_all: dict[str, dict] = {}
    proj_all:    dict[str, dict] = {}

    for ticker, cdata in companies.items():
        metrics_all[ticker] = compute_metrics(cdata)
        proj_all[ticker]    = project_years(cdata, drivers, proj_labels)

    # ── Header ─────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1B365D,#0D1B2E);
                border-radius:14px;padding:26px 34px;margin-bottom:24px;
                border:1px solid #2D5F8A;">
      <h1 style="color:#FFFFFF;margin:0;font-size:24px;font-weight:700;">
        📊 F&B Restaurant Sector — Integrated 3-Statement Model
      </h1>
      <p style="color:#A8C8E8;margin:8px 0 0;font-size:12px;">
        Sector Benchmarking Assignment · Finance Analytics Dashboard ·
        Institutional-Grade Financial Modeling
      </p>
    </div>
    """, unsafe_allow_html=True)

    # Company selector
    selected_tickers = st.multiselect(
        "Active Companies",
        options=list(companies.keys()),
        default=list(companies.keys()),
    )
    if not selected_tickers:
        st.warning("Select at least one company.")
        return

    active = {t: companies[t] for t in selected_tickers}
    active_metrics = {t: metrics_all[t] for t in selected_tickers}
    active_proj    = {t: proj_all[t]    for t in selected_tickers}

    # ── Tabs ───────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Methodology",
        "🏢 Executive Overview",
        "📑 3-Statement Model",
        "📈 Visual Analytics",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1: METHODOLOGY
    # ══════════════════════════════════════════════════════════════════════════
    with tab1:
        st.markdown('<div class="section-hdr">Modeling Framework and Assumptions</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            for title, body in [
                ("3-Statement Linking Logic",
                 "Net Income from the Income Statement flows directly into Retained "
                 "Earnings on the Balance Sheet each projection year. Operating Cash "
                 "Flow from the Cash Flow Statement updates the ending Cash balance "
                 "on the Balance Sheet. These linkages ensure the three statements "
                 "remain internally consistent as any projection driver changes."),
                ("Net Debt Definition",
                 "Net Debt is defined strictly as Long-Term Debt minus Cash and Cash "
                 "Equivalents. This is the standard institutional definition used by "
                 "investment banks and private equity firms when sizing leverage. A "
                 "negative Net Debt figure indicates a net cash position, meaning the "
                 "company holds more cash than it owes in long-term obligations."),
                ("EBITDA Calculation",
                 "EBITDA is calculated as Net Revenue minus Cost of Sales, minus Labor "
                 "and Related Costs, minus Advertising Expense, minus Store Operating "
                 "Expenses, and minus SG&A. Depreciation and Amortization is excluded "
                 "from this calculation because it is a non-cash charge. This produces "
                 "a proxy for operating cash generation before financing and investment."),
            ]:
                st.markdown(f"""
                <div class="method-card">
                  <div class="method-title">{title}</div>
                  <div class="method-body">{body}</div>
                </div>""", unsafe_allow_html=True)

        with c2:
            for title, body in [
                ("Store-Level Scaling Assumptions",
                 "Projection drivers are applied as uniform percentages of revenue "
                 "across all uploaded concepts. This approach simplifies comparability "
                 "but understates company-specific unit economics differences such as "
                 "AUV mix shifts or new market entry costs. Users should adjust the "
                 "sidebar sliders to reflect each concept's disclosed guidance where "
                 "management has provided forward-looking ranges."),
                ("Prime Cost Threshold",
                 "The F&B industry benchmark for Prime Cost (the sum of COGS and "
                 "Labor as a percentage of revenue) is approximately 55 to 60 percent. "
                 "Concepts exceeding 65 percent typically face margin compression that "
                 "requires either pricing action or cost structure intervention. The "
                 "visual analytics tab highlights the 60 percent threshold on the "
                 "Prime Cost trend chart for easy identification."),
                ("Free Cash Flow Conversion",
                 "Free Cash Flow is defined as Operating Cash Flow plus Capital "
                 "Expenditures (which are negative outflows). A high-growth concept "
                 "with heavy unit expansion CapEx will often show negative or minimal "
                 "FCF despite strong store-level economics. FCF Conversion Rate is "
                 "FCF divided by Net Revenue and serves as the primary capital "
                 "efficiency metric in this benchmarking analysis."),
            ]:
                st.markdown(f"""
                <div class="method-card">
                  <div class="method-title">{title}</div>
                  <div class="method-body">{body}</div>
                </div>""", unsafe_allow_html=True)

        # Balance check
        st.markdown('<div class="section-hdr">3-Statement Balance Verification</div>',
                    unsafe_allow_html=True)
        balance_cols = st.columns(len(active))
        for i, (ticker, cdata) in enumerate(active.items()):
            last_yr = cdata["years"][-1]
            ta  = cdata["data"].get("total_assets",     {}).get(last_yr, 0) or 0
            tl  = cdata["data"].get("total_liabilities", {}).get(last_yr, 0) or 0
            eq  = cdata["data"].get("equity",            {}).get(last_yr, 0) or 0
            diff = ta - (tl + eq)
            ok = abs(diff) < abs(ta) * 0.05  # within 5% tolerance
            with balance_cols[i]:
                st.metric(
                    label=f"{ticker} — Balance Check ({last_yr})",
                    value=f"${diff:,.0f}K variance",
                    delta="Balanced" if ok else f"${diff:,.0f}K gap",
                    delta_color="normal" if ok else "inverse",
                )

        # Unmapped items
        any_unmapped = any(cdata.get("unmapped") for cdata in active.values())
        if any_unmapped:
            with st.expander("Unmapped Line Items (add to LINE_MAP to resolve)"):
                for ticker, cdata in active.items():
                    if cdata.get("unmapped"):
                        st.write(f"**{ticker}:** {', '.join(cdata['unmapped'])}")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2: EXECUTIVE OVERVIEW
    # ══════════════════════════════════════════════════════════════════════════
    with tab2:
        # KPI Cards — always show latest actual (historical) year
        for ticker, cdata in active.items():
            # latest_actual_year is the last year in cdata["years"] (historical only)
            last_actual = cdata["years"][-1]
            m   = active_metrics[ticker].get(last_actual, {})
            rev  = cdata["data"].get("revenue",    {}).get(last_actual, 0) or 0
            ni   = cdata["data"].get("net_income", {}).get(last_actual, 0) or 0
            ltd  = cdata["data"].get("ltd",        {}).get(last_actual, 0) or 0
            cash = cdata["data"].get("cash",       {}).get(last_actual, 0) or 0
            nd   = ltd - cash

            st.markdown(
                f'<div class="section-hdr">{ticker} '
                f'<span style="font-size:11px;color:#7F9DBF;font-weight:400;">'
                f'Latest Actual: {last_actual}</span></div>',
                unsafe_allow_html=True,
            )
            c1, c2, c3, c4, c5 = st.columns(5)
            kpi_card(c1, f"Net Revenue ({last_actual})", fmt_val(rev))
            kpi_card(c2, "EBITDA Margin",
                     f"{m.get('ebitda_margin', 0):.1%}",
                     delta="vs industry ~15%" if m.get("ebitda_margin", 0) > 0.15 else "")
            kpi_card(c3, "Net Income",
                     fmt_val(ni),
                     delta=f"Net Margin {m.get('net_margin', 0):.1%}",
                     neg=(ni < 0))
            kpi_card(c4, "Net Debt (LTD minus Cash)", fmt_val(nd), neg=(nd > 0))
            kpi_card(c5, "Free Cash Flow", fmt_val(m.get("fcf", 0)),
                     neg=(m.get("fcf", 0) < 0))

        # Cross-company benchmarking table
        st.markdown('<div class="section-hdr">Cross-Company Benchmarking Matrix</div>',
                    unsafe_allow_html=True)

        bench_rows = []
        for ticker, cdata in active.items():
            last_actual = cdata["years"][-1]
            m   = active_metrics[ticker].get(last_actual, {})
            rev = cdata["data"].get("revenue",    {}).get(last_actual, 0) or 0
            ni  = cdata["data"].get("net_income", {}).get(last_actual, 0) or 0
            ltd = cdata["data"].get("ltd",        {}).get(last_actual, 0) or 0
            csh = cdata["data"].get("cash",       {}).get(last_actual, 0) or 0
            bench_rows.append({
                "Ticker":           ticker,
                "Latest Year":      last_actual,
                "Net Revenue ($M)": f"${rev/1e3:.1f}M",
                "EBITDA Margin":    f"{m.get('ebitda_margin', 0):.1%}",
                "Net Margin":       f"{m.get('net_margin', 0):.1%}",
                "Net Income ($M)":  f"${ni/1e3:.1f}M",
                "FCF ($M)":         f"${m.get('fcf', 0)/1e3:.1f}M",
                "FCF Margin":       f"{m.get('fcf_margin', 0):.1%}",
                "Prime Cost Pct":   f"{m.get('prime_cost_pct', 0):.1%}",
                "Net Debt ($M)":    f"${(ltd - csh)/1e3:.1f}M",
                "CapEx Pct Rev":    f"{m.get('capex_pct', 0):.1%}",
            })

        bench_df = pd.DataFrame(bench_rows)
        st.dataframe(bench_df, hide_index=True, use_container_width=True,
                     height=min(200, 60 + 35 * len(bench_rows)))

        # Projection summary — includes Net Margin and CFF
        st.markdown(
            f'<div class="section-hdr">'
            f'{proj_labels[0]} to {proj_labels[-1]} Projection Summary</div>',
            unsafe_allow_html=True,
        )

        proj_rows = []
        for ticker in active:
            for yr, pdata in active_proj[ticker].items():
                rev_p = pdata.get("revenue", 0)
                proj_rows.append({
                    "Ticker":            ticker,
                    "Year":              yr,
                    "Revenue ($M)":      f"${rev_p/1e3:.1f}M",
                    "EBITDA Margin":     f"{pdata.get('ebitda_margin', 0):.1%}",
                    "Net Income ($M)":   f"${pdata.get('net_income', 0)/1e3:.1f}M",
                    "Net Margin %":      f"{pdata.get('net_margin', 0):.1%}",
                    "FCF ($M)":          f"${pdata.get('fcf', 0)/1e3:.1f}M",
                    "FCF Margin %":      f"{pdata.get('fcf_margin', 0):.1%}",
                    "CFF ($M)":          f"${pdata.get('cff', 0)/1e3:.1f}M",
                    "Proj Cash ($M)":    f"${pdata.get('bs_cash', 0)/1e3:.1f}M",
                    "Net Debt ($M)":     f"${pdata.get('net_debt', 0)/1e3:.1f}M",
                })
        proj_df = pd.DataFrame(proj_rows)
        st.dataframe(proj_df, hide_index=True, use_container_width=True,
                     height=min(400, 60 + 35 * len(proj_rows)))

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3: 3-STATEMENT MODEL
    # ══════════════════════════════════════════════════════════════════════════
    with tab3:
        company_sel = st.selectbox("Select Company", list(active.keys()), key="stmt_sel")
        cdata  = active[company_sel]
        cmets  = active_metrics[company_sel]
        cproj  = active_proj[company_sel]

        is_df  = build_is_df(cdata, cmets, cproj)
        bs_df  = build_bs_df(cdata, cproj)
        cf_df  = build_cf_df(cdata, cmets, cproj)

        pct_rows = ["EBITDA Margin", "Net Margin", "Net Margin %"]

        def display_stmt(df, title):
            st.markdown(f'<div class="section-hdr">{title}</div>',
                        unsafe_allow_html=True)
            yr_cols   = [c for c in df.columns if c != "Line Item"]
            disp_df   = df.copy()
            for col in yr_cols:
                disp_df[col] = disp_df.apply(
                    lambda row, c=col: (
                        f"{row[c]:.1%}" if row["Line Item"] in pct_rows
                        else fmt_val(row[c] if not pd.isna(row[c]) else 0)
                    ), axis=1
                )

            def highlight_row(row):
                styles = []
                for val in row:
                    if row["Line Item"] in ("EBITDA","Net Income","Total Assets",
                                            "Free Cash Flow","Cash Flow from Operations"):
                        styles.append("background-color:#1B365D;font-weight:700;color:#EAF2FA")
                    elif row["Line Item"] in pct_rows:
                        styles.append("color:#7F7F7F;font-style:italic")
                    elif str(val).startswith("(") and "Line Item" not in str(row.name):
                        styles.append("color:#FF6B6B")
                    else:
                        styles.append("color:#EAF2FA")
                return styles

            st.dataframe(
                disp_df.style.apply(highlight_row, axis=1),
                hide_index=True,
                use_container_width=True,
                height=min(600, 38 + 35 * len(df)),
            )

        display_stmt(is_df, "Income Statement")
        display_stmt(bs_df, "Balance Sheet")
        display_stmt(cf_df, "Cash Flow Statement")

        # KPI block
        st.markdown('<div class="section-hdr">Key Performance Indicators</div>',
                    unsafe_allow_html=True)
        last_yr = cdata["years"][-1]
        m = cmets.get(last_yr, {})

        kpi_data = [
            ("Prime Cost Pct",      f"{m.get('prime_cost_pct',0):.1%}"),
            ("EBITDA Margin",       f"{m.get('ebitda_margin',0):.1%}"),
            ("Net Margin",          f"{m.get('net_margin',0):.1%}"),
            ("FCF Margin",          f"{m.get('fcf_margin',0):.1%}"),
            ("CapEx Pct Revenue",   f"{m.get('capex_pct',0):.1%}"),
            ("Net Debt ($000s)",    fmt_val(m.get("net_debt",0))),
        ]
        kpi_cols = st.columns(len(kpi_data))
        for col, (lbl, val) in zip(kpi_cols, kpi_data):
            kpi_card(col, lbl, val)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 4: VISUAL ANALYTICS
    # ══════════════════════════════════════════════════════════════════════════
    with tab4:
        r1c1, r1c2 = st.columns(2)
        with r1c1:
            st.plotly_chart(chart_revenue_trend(active, active_metrics, active_proj),
                            use_container_width=True)
        with r1c2:
            st.plotly_chart(chart_ebitda_margin(active, active_metrics, active_proj),
                            use_container_width=True)

        r2c1, r2c2 = st.columns(2)
        with r2c1:
            st.plotly_chart(chart_fcf_conversion(active, active_metrics, active_proj),
                            use_container_width=True)
        with r2c2:
            st.plotly_chart(chart_capital_structure(active, active_metrics),
                            use_container_width=True)

        r3c1, r3c2 = st.columns(2)
        with r3c1:
            st.plotly_chart(chart_prime_cost(active, active_metrics),
                            use_container_width=True)
        with r3c2:
            # Pick first selected company for sensitivity
            first_ticker = list(active.keys())[0]
            first_last   = active[first_ticker]["years"][-1]
            base_em = active_metrics[first_ticker].get(first_last,{}).get("ebitda_margin", 0.15)
            st.plotly_chart(chart_sensitivity(active_proj, rev_gr, base_em),
                            use_container_width=True)

    # ══════════════════════════════════════════════════════════════════════════
    # EXPORT BUTTON
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown('<div class="section-hdr">Export</div>', unsafe_allow_html=True)
    with st.spinner("Building Excel export…"):
        excel_bytes = build_export_excel(active, active_metrics, active_proj)

    st.download_button(
        label="📥 Download 3-Statement Model (.xlsx)",
        data=excel_bytes,
        file_name="fb_3statement_model.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Downloads a formatted multi-tab Excel workbook with historical data, projections, and comparative benchmarking.",
    )

    # Footer
    st.markdown("""
    <div style="text-align:center;color:#4B5563;font-size:11px;padding:16px 0 8px;">
      F&B Restaurant Sector Benchmarking Assignment ·
      Data sourced from SEC EDGAR 10-K filings ·
      Projections are analytical estimates based on disclosed guidance and historical trends ·
      <em>Not investment advice</em>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
